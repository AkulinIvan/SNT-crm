from typing import List
import logging
import traceback
import os
import tempfile

from django.db import models, DatabaseError, IntegrityError
from django.core.exceptions import ValidationError, PermissionDenied
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.shortcuts import render
from django.views import View
from django.db.models import Count, Q, Avg, Sum
from django.db import transaction
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from rest_framework import permissions
from django.http import JsonResponse, HttpResponse

from .services import rosreestr_service
from common.mixins import OrganizationMixin
from accounts.permissions import IsManagerOrAbove
from subscriptions.decorators import subscription_required
from .models import LandPlot
from .serializers import (
    LandPlotListSerializer,
    LandPlotDetailSerializer,
    LandPlotGeoSerializer,
)
from .excel_importer import ExcelImporter

logger = logging.getLogger(__name__)


class LandPlotViewSet(OrganizationMixin, viewsets.ModelViewSet):
    """
    ViewSet для управления земельными участками.
    
    Endpoints:
    - GET /api/plots/ - список участков с пагинацией и статистикой
    - POST /api/plots/ - создание участка
    - GET /api/plots/{id}/ - детали участка
    - PUT/PATCH /api/plots/{id}/ - обновление участка
    - DELETE /api/plots/{id}/ - удаление участка
    - GET /api/plots/geo/ - геоданные для карты
    - GET /api/plots/stats/ - статистика по участкам
    - POST /api/plots/bulk-update-status/ - массовое обновление статусов
    - POST /api/plots/import-excel/ - импорт из Excel
    """
    queryset = LandPlot.objects.prefetch_related('ownerships__owner')
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status']
    search_fields = ['plot_number', 'cadastral_number', 'address']
    ordering_fields = ['plot_number', 'area_sqm', 'cadastral_number', 'created_at', 'status']
    ordering = ['plot_number']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._logger = logging.getLogger(f'{__name__}.LandPlotViewSet')

    def get_permissions(self):
        """Определение прав доступа в зависимости от действия"""
        try:
            if self.action in ('create', 'update', 'partial_update', 'destroy', 
                              'deactivate', 'activate', 'bulk-update-status',
                              'bulk-set-coordinates', 'import-excel', 
                              'save-boundaries', 'set-coordinates'):
                return [permissions.IsAuthenticated(), IsManagerOrAbove()]
            return [permissions.IsAuthenticated()]
        except Exception as e:
            self._logger.error(f"Error determining permissions: {e}", exc_info=True)
            return [permissions.IsAuthenticated()]

    def get_serializer_class(self):
        """Выбор сериализатора в зависимости от действия"""
        try:
            if self.action == 'list':
                return LandPlotListSerializer
            elif self.action == 'geo':
                return LandPlotGeoSerializer
            return LandPlotDetailSerializer
        except Exception as e:
            self._logger.error(f"Error selecting serializer: {e}", exc_info=True)
            return LandPlotDetailSerializer

    def get_queryset(self):
        """Расширенная фильтрация с учетом организации"""
        try:
            queryset = super().get_queryset()
            
            # Фильтрация по организации
            if hasattr(self.request, 'current_organization') and self.request.current_organization:
                queryset = queryset.filter(organization=self.request.current_organization)
                self._logger.debug(
                    f"Filtered by organization: {self.request.current_organization.short_name}"
                )
            
            # Применяем фильтры из параметров запроса
            queryset = self._apply_query_filters(queryset)
            
            # Применяем distinct для избежания дублей при join'ах
            queryset = queryset.distinct()
            
            self._logger.debug(f"Final queryset count: {queryset.count()}")
            return queryset
            
        except DatabaseError as e:
            self._logger.error(f"Database error in get_queryset: {e}", exc_info=True)
            return LandPlot.objects.none()
        except Exception as e:
            self._logger.error(f"Critical error in get_queryset: {e}", exc_info=True)
            return LandPlot.objects.none()

    def _apply_query_filters(self, queryset):
        """Применение всех фильтров из параметров запроса"""
        params = self.request.query_params
        
        # 1. Фильтр по статусу
        status_filter = params.get('status')
        if status_filter and status_filter in ['active', 'abandoned', 'disputed']:
            queryset = queryset.filter(status=status_filter)
            self._logger.debug(f"Applied status filter: {status_filter}")
        
        # 2. Фильтр по координатам
        has_coordinates = params.get('has_coordinates')
        if has_coordinates is not None:
            if has_coordinates.lower() == 'true':
                queryset = queryset.filter(
                    latitude__isnull=False, 
                    longitude__isnull=False
                )
            elif has_coordinates.lower() == 'false':
                queryset = queryset.filter(
                    Q(latitude__isnull=True) | Q(longitude__isnull=True)
                )
            self._logger.debug(f"Applied has_coordinates filter: {has_coordinates}")
        
        # 3. Фильтр по границам
        has_boundaries = params.get('has_boundaries')
        if has_boundaries is not None:
            if has_boundaries.lower() == 'true':
                queryset = queryset.filter(
                    boundaries__isnull=False
                ).exclude(boundaries=[])
            elif has_boundaries.lower() == 'false':
                queryset = queryset.filter(
                    Q(boundaries__isnull=True) | Q(boundaries=[])
                )
            self._logger.debug(f"Applied has_boundaries filter: {has_boundaries}")
        
        # 4. Фильтр по наличию владельцев
        has_owners = params.get('has_owners')
        if has_owners is not None:
            if has_owners.lower() == 'true':
                queryset = queryset.annotate(
                    owners_cnt=Count('ownerships')
                ).filter(owners_cnt__gt=0)
            elif has_owners.lower() == 'false':
                queryset = queryset.annotate(
                    owners_cnt=Count('ownerships')
                ).filter(owners_cnt=0)
            self._logger.debug(f"Applied has_owners filter: {has_owners}")
        
        # 5. Фильтр по площади
        area_min = params.get('area_min')
        if area_min:
            try:
                queryset = queryset.filter(area_sqm__gte=float(area_min))
                self._logger.debug(f"Applied area_min filter: {area_min}")
            except ValueError:
                self._logger.warning(f"Invalid area_min value: {area_min}")
            
        area_max = params.get('area_max')
        if area_max:
            try:
                queryset = queryset.filter(area_sqm__lte=float(area_max))
                self._logger.debug(f"Applied area_max filter: {area_max}")
            except ValueError:
                self._logger.warning(f"Invalid area_max value: {area_max}")
            
        # 6. Поиск по тексту
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(plot_number__icontains=search) |
                Q(cadastral_number__icontains=search) |
                Q(address__icontains=search)
            )
            self._logger.debug(f"Applied search filter: '{search}'")
        
        return queryset

    def perform_create(self, serializer):
        """При создании автоматически подставляем организацию"""
        try:
            if hasattr(self.request, 'current_organization') and self.request.current_organization:
                plot = serializer.save(organization=self.request.current_organization)
                self._logger.info(
                    f"Plot created: #{plot.plot_number} (ID: {plot.id}) "
                    f"in organization: {self.request.current_organization.short_name}"
                )
            else:
                plot = serializer.save()
                self._logger.info(f"Plot created without organization: #{plot.plot_number} (ID: {plot.id})")
        except Exception as e:
            self._logger.error(f"Error in perform_create: {e}", exc_info=True)
            raise

    def list(self, request, *args, **kwargs):
        """Расширенный list с дополнительной статистикой"""
        self._logger.info(
            f"Listing plots for user: {request.user.username} "
            f"(org: {getattr(request, 'current_organization', None)})"
        )
        
        try:
            queryset = self.filter_queryset(self.get_queryset())

            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                response = self.get_paginated_response(serializer.data)

                # Вычисляем статистику
                try:
                    total_area = queryset.aggregate(total=Sum('area_sqm'))['total'] or 0
                    
                    response.data['stats'] = {
                        'total': queryset.count(),
                        'total_area': round(total_area, 2),
                        'by_status': {
                            'active': queryset.filter(status='active').count(),
                            'abandoned': queryset.filter(status='abandoned').count(),
                            'disputed': queryset.filter(status='disputed').count(),
                        }
                    }
                    self._logger.debug(f"Stats calculated: {response.data['stats']['total']} plots")
                except Exception as e:
                    self._logger.error(f"Error calculating stats: {e}", exc_info=True)
                
                return response

            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
            
        except DatabaseError as e:
            self._logger.error(f"Database error in list: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка базы данных при получении списка участков'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            self._logger.error(f"Critical error in list: {e}", exc_info=True)
            return Response(
                {'detail': 'Внутренняя ошибка сервера'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def create(self, request, *args, **kwargs):
        """Создание участка с проверкой лимитов тарифа"""
        plot_number = request.data.get('plot_number', 'unknown')
        self._logger.info(f"Creating plot: {plot_number} by user: {request.user.username}")
        
        try:
            # Проверка лимитов тарифа
            organization = request.current_organization
            if organization:
                try:
                    is_allowed, current, max_limit, message = organization.check_tariff_limit('plots')
                    
                    if not is_allowed:
                        self._logger.warning(
                            f"Tariff limit reached for {organization.short_name}: "
                            f"plots {current}/{max_limit}"
                        )
                        return Response(
                            {
                                'detail': message,
                                'code': 'tariff_limit_reached',
                                'current': current,
                                'max': max_limit,
                                'tariff': (
                                    organization.subscription.tariff.name 
                                    if hasattr(organization, 'subscription') and organization.subscription 
                                    else None
                                )
                            },
                            status=status.HTTP_403_FORBIDDEN
                        )
                except Exception as e:
                    # Логируем ошибку, но не блокируем создание
                    self._logger.error(f"Error checking tariff limits: {e}", exc_info=True)
            
            return super().create(request, *args, **kwargs)
            
        except IntegrityError as e:
            self._logger.error(f"Integrity error creating plot: {e}", exc_info=True)
            return Response(
                {'detail': 'Участок с таким кадастровым номером уже существует'},
                status=status.HTTP_409_CONFLICT
            )
        except ValidationError as e:
            self._logger.warning(f"Validation error creating plot: {e}")
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            self._logger.error(f"Error creating plot: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при создании участка'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def destroy(self, request, *args, **kwargs):
        """Безопасное удаление с проверками"""
        land_plot = self.get_object()
        
        self._logger.warning(
            f"Attempting to delete plot #{land_plot.plot_number} "
            f"(ID: {land_plot.id}) by user: {request.user.username}"
        )
        
        try:
            # Проверяем наличие активных начислений
            has_active_assessments = self._check_active_assessments(land_plot)
            
            if has_active_assessments:
                self._logger.warning(
                    f"Cannot delete plot #{land_plot.plot_number}: has active assessments"
                )
                return Response(
                    {
                        'detail': 'Невозможно удалить участок с неоплаченными начислениями.',
                        'code': 'has_active_assessments'
                    },
                    status=status.HTTP_409_CONFLICT
                )
            
            owners_count = land_plot.ownerships.count()
            
            with transaction.atomic():
                land_plot.delete()
                
                self._logger.info(
                    f"Plot #{land_plot.plot_number} deleted successfully "
                    f"(had {owners_count} owners)"
                )
                
                return Response(status=status.HTTP_204_NO_CONTENT)
                
        except IntegrityError as e:
            self._logger.error(f"Integrity error deleting plot: {e}", exc_info=True)
            return Response(
                {'detail': 'Невозможно удалить участок: есть связанные данные'},
                status=status.HTTP_409_CONFLICT
            )
        except Exception as e:
            self._logger.error(f"Error deleting plot: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при удалении участка'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _check_active_assessments(self, land_plot):
        """Проверка наличия активных начислений на участок"""
        try:
            from payments.models import Assessment
            return Assessment.objects.filter(
                land_plot=land_plot,
                status__in=['pending', 'partial', 'overdue']
            ).exists()
        except ImportError:
            self._logger.debug("Payments module not available for assessment check")
            return False
        except Exception as e:
            self._logger.error(f"Error checking assessments: {e}", exc_info=True)
            return False

    @action(detail=False, methods=['get'], url_path='geo')
    def geo(self, request):
        """
        GET /api/plots/geo/
        Возвращает координаты и границы участков для карты.
        """
        self._logger.info(
            f"Geo data requested by user: {request.user.username}, "
            f"params: {dict(request.query_params)}"
        )
        
        try:
            status_filter = request.query_params.get('status')
            has_owners = request.query_params.get('has_owners')
            show_without_coords = request.query_params.get('show_without_coords', 'false').lower() == 'true'

            # Базовый queryset
            if show_without_coords:
                queryset = LandPlot.objects.all()
                self._logger.debug("Showing all plots (including without coordinates)")
            else:
                queryset = LandPlot.objects.filter(
                    Q(latitude__isnull=False, longitude__isnull=False) |
                    Q(boundaries__isnull=False)
                ).exclude(boundaries=[])

            # Фильтрация по организации
            if not request.user.is_superuser and not request.user.is_admin:
                if hasattr(request, 'current_organization') and request.current_organization:
                    queryset = queryset.filter(organization=request.current_organization)
                else:
                    self._logger.debug("No organization for user, returning empty geo data")
                    return Response({'count': 0, 'results': []})

            # Применяем фильтры
            if status_filter:
                queryset = queryset.filter(status=status_filter)

            if has_owners is not None:
                queryset = queryset.annotate(owners_count=Count('ownerships'))
                if has_owners.lower() == 'true':
                    queryset = queryset.filter(owners_count__gt=0)
                else:
                    queryset = queryset.filter(owners_count=0)

            queryset = queryset.prefetch_related('ownerships__owner')

            count = queryset.count()
            self._logger.info(f"Geo API: found {count} plots")

            serializer = LandPlotGeoSerializer(queryset, many=True)
            return Response({
                'count': count,
                'results': serializer.data
            })
            
        except DatabaseError as e:
            self._logger.error(f"Database error in geo: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка базы данных'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            self._logger.error(f"Error in geo endpoint: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при получении геоданных'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='set-coordinates')
    def set_coordinates(self, request, pk=None):
        """
        POST /api/plots/{id}/set-coordinates/
        Быстрое обновление координат одного участка.
        """
        land_plot = self.get_object()
        
        self._logger.info(
            f"Setting coordinates for plot #{land_plot.plot_number} (ID: {land_plot.id})"
        )
        
        try:
            lat = request.data.get('latitude')
            lon = request.data.get('longitude')
            
            if lat is None or lon is None:
                return Response(
                    {'detail': 'Необходимо передать latitude и longitude'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            lat = float(lat)
            lon = float(lon)
            
            # Валидация диапазонов
            if not (-90 <= lat <= 90):
                raise ValueError('Широта должна быть от -90 до 90')
            if not (-180 <= lon <= 180):
                raise ValueError('Долгота должна быть от -180 до 180')
            
            land_plot.latitude = lat
            land_plot.longitude = lon
            land_plot.full_clean()
            land_plot.save(update_fields=['latitude', 'longitude', 'updated_at'])
            
            self._logger.info(
                f"Coordinates updated for plot #{land_plot.plot_number}: "
                f"lat={lat:.6f}, lon={lon:.6f}"
            )
            
            serializer = LandPlotDetailSerializer(land_plot)
            return Response(serializer.data)
            
        except ValueError as e:
            self._logger.warning(f"Invalid coordinates: {e}")
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except ValidationError as e:
            self._logger.warning(f"Validation error: {e}")
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            self._logger.error(f"Error setting coordinates: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при обновлении координат'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='load-boundaries')
    def load_boundaries(self, request, pk=None):
        """
        POST /api/plots/{id}/load-boundaries/
        Загрузка границ участка из Росреестра.
        """
        land_plot = self.get_object()
        
        self._logger.info(
            f"Loading boundaries for plot #{land_plot.plot_number} "
            f"(cadastral: {land_plot.cadastral_number})"
        )

        if not land_plot.cadastral_number:
            self._logger.warning(f"No cadastral number for plot #{land_plot.plot_number}")
            return Response(
                {'detail': 'Отсутствует кадастровый номер'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            start_time = timezone.now()
            
            boundaries = rosreestr_service.get_parcel_boundaries(
                land_plot.cadastral_number
            )

            elapsed = (timezone.now() - start_time).total_seconds()

            if boundaries and len(boundaries) >= 3:
                land_plot.boundaries = boundaries
                land_plot.rosreestr_updated = timezone.now()
                land_plot.save(update_fields=['boundaries', 'rosreestr_updated', 'updated_at'])

                self._logger.info(
                    f"Boundaries loaded for plot #{land_plot.plot_number}: "
                    f"{len(boundaries)} points ({elapsed:.2f}s)"
                )

                return Response({
                    'detail': f'Границы загружены ({len(boundaries)} точек)',
                    'boundaries': boundaries,
                    'updated_at': land_plot.rosreestr_updated,
                })
            else:
                self._logger.warning(
                    f"Boundaries not found for plot #{land_plot.plot_number} ({elapsed:.2f}s)"
                )
                return Response(
                    {'detail': 'Границы не найдены в кадастре. Проверьте правильность кадастрового номера.'},
                    status=status.HTTP_404_NOT_FOUND
                )

        except Exception as e:
            self._logger.error(f"Error loading boundaries: {e}", exc_info=True)
            return Response(
                {'detail': f'Ошибка загрузки: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='bulk-load-boundaries')
    def bulk_load_boundaries(self, request):
        """
        POST /api/plots/bulk-load-boundaries/
        Массовая загрузка границ.
        """
        self._logger.info(f"Bulk load boundaries requested by user: {request.user.username}")
        
        try:
            plot_ids = request.data.get('plot_ids')
            load_all = request.data.get('load_all', False)
            delay = float(request.data.get('delay', 0.5))
            
            if load_all:
                plots = LandPlot.objects.filter(
                    cadastral_number__isnull=False
                ).exclude(cadastral_number='')
                self._logger.info(f"Bulk load: all plots ({plots.count()})")
            elif plot_ids:
                plots = LandPlot.objects.filter(id__in=plot_ids)
                self._logger.info(f"Bulk load: {len(plot_ids)} plots")
            else:
                return Response(
                    {'detail': 'Укажите plot_ids или load_all=true'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            results = rosreestr_service.batch_load_boundaries(plots, delay=delay)
            
            self._logger.info(
                f"Bulk load completed: {results['success']} success, "
                f"{results['failed']} failed, {results.get('skipped', 0)} skipped"
            )
            
            return Response(results)
            
        except Exception as e:
            self._logger.error(f"Error in bulk load: {e}", exc_info=True)
            return Response(
                {'detail': f'Ошибка массовой загрузки: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='check-rosreestr')
    def check_rosreestr(self, request, pk=None):
        """
        GET /api/plots/{id}/check-rosreestr/
        Проверка наличия данных в Росреестре без сохранения.
        """
        land_plot = self.get_object()
        
        self._logger.info(
            f"Checking Rosreestr for plot #{land_plot.plot_number} "
            f"(cad: {land_plot.cadastral_number})"
        )
        
        try:
            result = {
                'plot_id': land_plot.id,
                'plot_number': land_plot.plot_number,
                'cadastral_number': land_plot.cadastral_number,
                'has_boundaries_in_db': land_plot.has_boundaries,
                'checks': {}
            }
            
            # Проверка по кадастровому номеру
            if land_plot.cadastral_number:
                try:
                    boundaries = rosreestr_service.get_parcel_boundaries(land_plot.cadastral_number)
                    result['checks']['by_cadastral'] = {
                        'found': boundaries is not None,
                        'points_count': len(boundaries) if boundaries else 0
                    }
                except Exception as e:
                    self._logger.warning(f"Error checking by cadastral: {e}")
                    result['checks']['by_cadastral'] = {
                        'found': False,
                        'error': str(e)[:200]
                    }
            
            # Проверка по координатам
            if land_plot.has_coordinates:
                try:
                    parcel_info = rosreestr_service.get_parcel_by_coordinates(
                        land_plot.latitude, 
                        land_plot.longitude
                    )
                    result['checks']['by_coordinates'] = {
                        'found': parcel_info is not None,
                        'cadastral_number': parcel_info.get('cadastral_number') if parcel_info else None,
                        'address': parcel_info.get('address') if parcel_info else None
                    }
                except Exception as e:
                    self._logger.warning(f"Error checking by coordinates: {e}")
                    result['checks']['by_coordinates'] = {
                        'found': False,
                        'error': str(e)[:200]
                    }
            
            return Response(result)
            
        except Exception as e:
            self._logger.error(f"Error in check-rosreestr: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при проверке Росреестра'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='boundaries')
    def get_boundaries(self, request, pk=None):
        """
        GET /api/plots/{id}/boundaries/
        Получение сохраненных границ участка.
        """
        land_plot = self.get_object()
        
        if not land_plot.has_boundaries:
            return Response(
                {'detail': 'Границы не загружены'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        return Response({
            'id': land_plot.id,
            'plot_number': land_plot.plot_number,
            'cadastral_number': land_plot.cadastral_number,
            'boundaries': land_plot.boundaries,
            'rosreestr_updated': land_plot.rosreestr_updated,
        })
    
    @action(detail=False, methods=['get'], url_path='all-boundaries')
    def all_boundaries(self, request):
        """
        GET /api/plots/all-boundaries/
        Получение всех загруженных границ для карты.
        """
        self._logger.info(f"All boundaries requested by user: {request.user.username}")
        
        try:
            plots = LandPlot.objects.filter(
                boundaries__isnull=False
            ).exclude(boundaries=[])
            
            data = []
            for plot in plots:
                data.append({
                    'id': plot.id,
                    'plot_number': plot.plot_number,
                    'cadastral_number': plot.cadastral_number,
                    'status': plot.status,
                    'boundaries': plot.boundaries,
                    'area_sqm': plot.area_sqm,
                    'owners_count': plot.owners_count,
                })
            
            self._logger.info(f"All boundaries: {len(data)} plots returned")
            
            return Response({
                'count': len(data),
                'results': data,
            })
            
        except Exception as e:
            self._logger.error(f"Error in all-boundaries: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при получении границ'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='save-boundaries')
    def save_boundaries(self, request, pk=None):
        """
        POST /api/plots/{id}/save-boundaries/
        Ручное сохранение границ участка с автоматической нормализацией.
        """
        land_plot = self.get_object()
        boundaries = request.data.get('boundaries', [])
        source = request.data.get('source', 'manual')
        
        self._logger.info(
            f"Saving boundaries for plot #{land_plot.plot_number}: "
            f"{len(boundaries)} points (source: {source})"
        )

        if not boundaries or len(boundaries) < 3:
            return Response(
                {'detail': 'Минимум 3 точки для построения полигона'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Валидация и нормализация координат
        validated = []
        invalid_count = 0
        for point in boundaries:
            if len(point) < 2:
                invalid_count += 1
                continue
            try:
                lat, lon = float(point[0]), float(point[1])
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    validated.append([lat, lon])
                else:
                    invalid_count += 1
            except (ValueError, TypeError):
                invalid_count += 1
                continue

        if invalid_count > 0:
            self._logger.warning(f"Skipped {invalid_count} invalid points")

        if len(validated) < 3:
            return Response(
                {'detail': 'Недостаточно валидных координат (минимум 3 точки)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Нормализация: удаляем дубликаты и замыкаем полигон
            normalized = self._normalize_boundaries(validated)

            land_plot.boundaries = normalized
            land_plot.rosreestr_updated = timezone.now()
            land_plot.save(update_fields=['boundaries', 'rosreestr_updated', 'updated_at'])

            self._logger.info(
                f"Boundaries saved for plot #{land_plot.plot_number}: "
                f"{len(normalized)} points (source: {source})"
            )

            return Response({
                'detail': f'Границы сохранены ({len(normalized)} точек)',
                'boundaries': normalized,
                'source': source,
                'normalized': len(normalized) != len(validated)
            })
            
        except Exception as e:
            self._logger.error(f"Error saving boundaries: {e}", exc_info=True)
            return Response(
                {'detail': f'Ошибка при сохранении границ: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _normalize_boundaries(self, boundaries: List) -> List:
        """Нормализация границ: замыкание и удаление дубликатов"""
        if not boundaries or len(boundaries) < 3:
            return boundaries
        
        # Удаляем дублирующиеся подряд точки
        unique = []
        for point in boundaries:
            if not unique or point != unique[-1]:
                unique.append(point)
        
        # Замыкаем полигон если нужно
        if len(unique) >= 3 and unique[0] != unique[-1]:
            unique.append(unique[0])
        
        return unique
    
    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        """Расширенная статистика по участкам с учетом СНТ пользователя."""
        self._logger.info(f"Stats requested by user: {request.user.username}")
        
        try:
            queryset = LandPlot.objects.all()

            # Фильтруем по организации
            if not request.user.is_superuser and not request.user.is_admin:
                if hasattr(request, 'current_organization') and request.current_organization:
                    queryset = queryset.filter(organization=request.current_organization)
                else:
                    return Response(self._get_empty_stats())

            total = queryset.count()
            
            if total == 0:
                return Response(self._get_empty_stats())

            # Подсчет участков с границами
            with_boundaries = queryset.filter(
                boundaries__isnull=False
            ).exclude(boundaries=[]).count()

            without_boundaries = queryset.filter(
                Q(boundaries__isnull=True) | Q(boundaries=[])
            ).count()

            stats = {
                'total': total,
                'active': queryset.filter(status='active').count(),
                'abandoned': queryset.filter(status='abandoned').count(),
                'disputed': queryset.filter(status='disputed').count(),
                'total_area': round(queryset.aggregate(total=Sum('area_sqm'))['total'] or 0, 2),
                'with_coordinates': queryset.filter(
                    latitude__isnull=False, 
                    longitude__isnull=False
                ).count(),
                'without_coordinates': queryset.filter(
                    Q(latitude__isnull=True) | Q(longitude__isnull=True)
                ).count(),
                'with_owners': queryset.annotate(
                    owners_count=Count('ownerships')
                ).filter(owners_count__gt=0).count(),
                'without_owners': queryset.annotate(
                    owners_count=Count('ownerships')
                ).filter(owners_count=0).count(),
                'average_area': round(
                    queryset.aggregate(avg=Avg('area_sqm'))['avg'] or 0, 2
                ),
                'average_owners_per_plot': round(
                    queryset.annotate(
                        owners_count=Count('ownerships')
                    ).aggregate(avg=Avg('owners_count'))['avg'] or 0, 1
                ),
                'last_added': queryset.order_by('-created_at').first().plot_number 
                    if queryset.exists() else None,
                'with_boundaries': with_boundaries,
                'without_boundaries': without_boundaries,
            }
            
            self._logger.info(f"Stats calculated: {total} plots, {with_boundaries} with boundaries")
            return Response(stats)
            
        except DatabaseError as e:
            self._logger.error(f"Database error in stats: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка базы данных при расчете статистики'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            self._logger.error(f"Error in stats: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при расчете статистики'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_empty_stats(self):
        """Пустая статистика"""
        return {
            'total': 0,
            'active': 0,
            'abandoned': 0,
            'disputed': 0,
            'total_area': 0,
            'with_coordinates': 0,
            'without_coordinates': 0,
            'with_owners': 0,
            'without_owners': 0,
            'average_area': 0,
            'average_owners_per_plot': 0,
            'last_added': None,
            'with_boundaries': 0,
            'without_boundaries': 0,
        }

    @action(detail=False, methods=['get'], url_path='near')
    def near_plots(self, request):
        """
        GET /api/plots/near/?lat=55.75&lon=37.62&radius=0.01
        Поиск участков рядом с указанными координатами.
        """
        self._logger.info(
            f"Near plots search: lat={request.query_params.get('lat')}, "
            f"lon={request.query_params.get('lon')}"
        )
        
        try:
            lat = request.query_params.get('lat')
            lon = request.query_params.get('lon')
            radius = float(request.query_params.get('radius', 0.01))
            
            if not lat or not lon:
                return Response(
                    {'detail': 'Укажите lat и lon'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            lat = float(lat)
            lon = float(lon)
            
            # Валидация координат
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                return Response(
                    {'detail': 'Координаты вне допустимого диапазона'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
        except ValueError:
            return Response(
                {'detail': 'Некорректные координаты'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        try:
            queryset = LandPlot.objects.filter(
                latitude__isnull=False,
                longitude__isnull=False,
                latitude__gte=lat - radius,
                latitude__lte=lat + radius,
                longitude__gte=lon - radius,
                longitude__lte=lon + radius,
            )
            
            count = queryset.count()
            self._logger.info(f"Near plots: found {count} plots within radius {radius}")
            
            serializer = LandPlotGeoSerializer(queryset, many=True)
            return Response({
                'center': {'lat': lat, 'lon': lon},
                'radius': radius,
                'count': count,
                'results': serializer.data
            })
            
        except Exception as e:
            self._logger.error(f"Error in near plots: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при поиске участков'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='export')
    def export_plots(self, request):
        """
        GET /api/plots/export/
        Экспорт списка участков в CSV.
        """
        self._logger.info(f"CSV export requested by user: {request.user.username}")
        
        try:
            import csv
            
            plots = self.filter_queryset(self.get_queryset())
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = (
                f'attachment; filename="land_plots_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            )
            response.write('\ufeff'.encode('utf8'))  # BOM для Excel
            
            writer = csv.writer(response)
            writer.writerow([
                'ID', 'Номер участка', 'Кадастровый номер', 'Площадь (м²)',
                'Адрес', 'Статус', 'Широта', 'Долгота', 
                'Кол-во владельцев', 'Примечания', 'Дата добавления'
            ])
            
            for plot in plots:
                writer.writerow([
                    plot.id,
                    plot.plot_number,
                    plot.cadastral_number,
                    plot.area_sqm,
                    plot.address or '',
                    plot.get_status_display(),
                    plot.latitude or '',
                    plot.longitude or '',
                    plot.ownerships.count(),
                    plot.notes or '',
                    plot.created_at.strftime('%d.%m.%Y'),
                ])
            
            self._logger.info(f"CSV export: {plots.count()} plots exported")
            return response
            
        except Exception as e:
            self._logger.error(f"Error in CSV export: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при экспорте данных'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """Импорт данных из Excel"""
        self._logger.info(f"Excel import started by user: {request.user.username}")
        
        file = request.FILES.get('excel_file')
        
        if not file:
            return Response(
                {'error': 'Файл не выбран'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not file.name.endswith(('.xlsx', '.xls')):
            self._logger.warning(f"Invalid file format: {file.name}")
            return Response(
                {'error': 'Поддерживаются только файлы Excel (.xlsx, .xls)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверка размера файла (макс 10 МБ)
        if file.size > 10 * 1024 * 1024:
            self._logger.warning(f"File too large: {file.size} bytes")
            return Response(
                {'error': 'Размер файла не должен превышать 10 МБ'},
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
            )
        
        organization_id = None
        if hasattr(request, 'current_organization') and request.current_organization:
            organization_id = request.current_organization.id
            self._logger.info(f"Importing to organization: {request.current_organization.short_name}")
        
        tmp_file_path = None
        
        try:
            # Сохраняем файл во временный
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in file.chunks():
                    tmp_file.write(chunk)
                tmp_file_path = tmp_file.name
            
            self._logger.debug(f"File saved to temp: {tmp_file_path}")
            
            # Импортируем данные
            importer = ExcelImporter(tmp_file_path, organization_id)
            stats = importer.import_data()
            
            self._logger.info(
                f"Import completed: {stats.get('total_rows', 0)} rows, "
                f"{stats.get('plots_created', 0)} plots created, "
                f"{stats.get('owners_created', 0)} owners created, "
                f"{stats.get('errors', 0)} errors"
            )
            
            return Response({
                'success': True,
                'stats': stats,
                'errors': importer.errors[:20],
                'warnings': importer.warnings[:20],
            })
            
        except MemoryError as e:
            self._logger.error(f"Memory error during import: {e}")
            return Response(
                {'error': 'Файл слишком большой для обработки'},
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
            )
        except Exception as e:
            self._logger.error(f"Import error: {e}", exc_info=True)
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            # Удаляем временный файл
            if tmp_file_path and os.path.exists(tmp_file_path):
                try:
                    os.unlink(tmp_file_path)
                    self._logger.debug(f"Temp file removed: {tmp_file_path}")
                except Exception as e:
                    self._logger.error(f"Error removing temp file: {e}")
    
    @action(detail=False, methods=['post'], url_path='bulk-update-status')
    def bulk_update_status(self, request):
        """
        POST /api/plots/bulk-update-status/
        Массовое обновление статусов.
        """
        self._logger.info(f"Bulk status update by user: {request.user.username}")
        
        try:
            plot_ids = request.data.get('plot_ids', [])
            new_status = request.data.get('status')
            
            if not plot_ids:
                return Response(
                    {'detail': 'Укажите plot_ids'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            valid_statuses = [s[0] for s in LandPlot.STATUS_CHOICES]
            if new_status not in valid_statuses:
                return Response(
                    {'detail': f'Недопустимый статус. Допустимые: {", ".join(valid_statuses)}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            with transaction.atomic():
                updated_count = LandPlot.objects.filter(
                    id__in=plot_ids
                ).update(
                    status=new_status,
                    updated_at=timezone.now()
                )
            
            self._logger.info(
                f"Bulk status update: {updated_count}/{len(plot_ids)} plots "
                f"set to '{new_status}'"
            )
            
            return Response({
                'detail': f'Обновлено участков: {updated_count}',
                'updated_count': updated_count,
                'requested_count': len(plot_ids),
                'new_status': new_status,
            })
            
        except DatabaseError as e:
            self._logger.error(f"Database error in bulk update: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка базы данных при обновлении'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            self._logger.error(f"Error in bulk update: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при обновлении'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='bulk-set-coordinates')
    def bulk_set_coordinates(self, request):
        """
        POST /api/plots/bulk-set-coordinates/
        Массовое обновление координат.
        """
        self._logger.info(f"Bulk coordinates update by user: {request.user.username}")
        
        coordinates_data = request.data
        
        if not isinstance(coordinates_data, list):
            return Response(
                {'detail': 'Ожидается массив объектов с id, latitude, longitude'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        updated = []
        errors = []
        
        try:
            with transaction.atomic():
                for item in coordinates_data:
                    try:
                        plot_id = item.get('id')
                        lat = float(item.get('latitude'))
                        lon = float(item.get('longitude'))
                        
                        if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                            errors.append({
                                'id': plot_id,
                                'error': 'Координаты вне допустимого диапазона'
                            })
                            continue
                        
                        plot = LandPlot.objects.filter(id=plot_id).first()
                        if not plot:
                            errors.append({
                                'id': plot_id,
                                'error': 'Участок не найден'
                            })
                            continue
                        
                        plot.latitude = lat
                        plot.longitude = lon
                        plot.save(update_fields=['latitude', 'longitude', 'updated_at'])
                        updated.append(plot_id)
                        
                    except ValueError as e:
                        errors.append({
                            'id': item.get('id'),
                            'error': f'Некорректные координаты: {str(e)}'
                        })
                    except Exception as e:
                        errors.append({
                            'id': item.get('id'),
                            'error': str(e)
                        })
            
            self._logger.info(
                f"Bulk coordinates: {len(updated)} updated, {len(errors)} errors"
            )
            
            return Response({
                'updated': len(updated),
                'errors': len(errors),
                'error_details': errors[:10],
            })
            
        except Exception as e:
            self._logger.error(f"Error in bulk coordinates: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при обновлении координат'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='tariff-warning')
    def tariff_warning(self, request):
        """Проверка приближения к лимитам тарифа"""
        self._logger.debug(f"Tariff warning check by user: {request.user.username}")
        
        try:
            organization = request.current_organization
            
            if not organization:
                return Response(
                    {'detail': 'Организация не найдена'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            subscription = getattr(organization, 'subscription', None)
            
            if not subscription or not subscription.is_active:
                return Response({
                    'has_subscription': False,
                    'warning': None,
                    'message': 'Нет активной подписки. Некоторые функции могут быть недоступны.'
                })
            
            tariff = subscription.tariff
            warnings = []
            
            # Проверка лимита владельцев
            owners_current = organization.owners_count
            owners_max = tariff.max_owners
            owners_percent = (owners_current / owners_max * 100) if owners_max > 0 else 0
            
            if owners_percent >= 90:
                warnings.append({
                    'type': 'owners',
                    'current': owners_current,
                    'max': owners_max,
                    'percent': round(owners_percent, 1),
                    'message': f'Достигнуто {owners_percent:.0f}% лимита владельцев ({owners_current}/{owners_max})'
                })
            
            # Проверка лимита участков
            plots_current = organization.plots_count
            plots_max = tariff.max_plots
            plots_percent = (plots_current / plots_max * 100) if plots_max > 0 else 0
            
            if plots_percent >= 90:
                warnings.append({
                    'type': 'plots',
                    'current': plots_current,
                    'max': plots_max,
                    'percent': round(plots_percent, 1),
                    'message': f'Достигнуто {plots_percent:.0f}% лимита участков ({plots_current}/{plots_max})'
                })
            
            # Проверка срока подписки
            days_left = subscription.days_left
            if days_left and days_left <= 30:
                warnings.append({
                    'type': 'subscription',
                    'days_left': days_left,
                    'message': f'Подписка истекает через {days_left} дней. Рекомендуем продлить.'
                })
            
            return Response({
                'has_subscription': True,
                'tariff': tariff.name,
                'warnings': warnings,
                'limits': {
                    'owners': {
                        'current': owners_current, 
                        'max': owners_max, 
                        'percent': round(owners_percent, 1)
                    },
                    'plots': {
                        'current': plots_current, 
                        'max': plots_max, 
                        'percent': round(plots_percent, 1)
                    },
                    'users': {
                        'current': organization.users_count, 
                        'max': tariff.max_users
                    }
                }
            })
            
        except Exception as e:
            self._logger.error(f"Error in tariff warning: {e}", exc_info=True)
            return Response(
                {'detail': 'Ошибка при проверке тарифа'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ==================== Веб-вьюхи ====================

class LandPlotListView(View):
    """Страница со списком участков."""
    _logger = logging.getLogger(f'{__name__}.LandPlotListView')
    
    @method_decorator(login_required)
    def get(self, request):
        try:
            return render(request, 'land/list.html', {'active_page': 'plots'})
        except Exception as e:
            self._logger.error(f"Error rendering plot list: {e}", exc_info=True)
            return render(request, 'error.html', {'error': 'Ошибка загрузки страницы'})


class LandPlotDetailView(View):
    """Страница карточки участка."""
    _logger = logging.getLogger(f'{__name__}.LandPlotDetailView')
    
    @method_decorator(login_required)
    def get(self, request, pk):
        try:
            self._logger.debug(f"Plot detail page requested: pk={pk}")
            return render(request, 'land/detail.html', {
                'active_page': 'plots',
                'plot_id': pk,
            })
        except Exception as e:
            self._logger.error(f"Error rendering plot detail: {e}", exc_info=True)
            return render(request, 'error.html', {'error': 'Ошибка загрузки страницы'})


class LandPlotMapView(View):
    """Страница с картой СНТ (требует подписки)"""
    _logger = logging.getLogger(f'{__name__}.LandPlotMapView')
    
    @method_decorator(login_required)
    @method_decorator(subscription_required(feature='map', redirect_url='subscription_plans'))
    def get(self, request):
        try:
            return render(request, 'land/map.html', {'active_page': 'map'})
        except Exception as e:
            self._logger.error(f"Error rendering map: {e}", exc_info=True)
            return render(request, 'error.html', {'error': 'Ошибка загрузки карты'})


class DashboardView(View):
    """Главная страница дашборда"""
    _logger = logging.getLogger(f'{__name__}.DashboardView')
    
    def get(self, request):
        try:
            return render(request, 'users/dashboard.html', {'active_page': 'dashboard'})
        except Exception as e:
            self._logger.error(f"Error rendering dashboard: {e}", exc_info=True)
            return render(request, 'error.html', {'error': 'Ошибка загрузки дашборда'})


class ExcelImportView(View):
    """Импорт данных из Excel"""
    _logger = logging.getLogger(f'{__name__}.ExcelImportView')
    
    @method_decorator(login_required)
    def get(self, request):
        return render(request, 'land/excel_import.html', {'active_page': 'import'})
    
    @method_decorator(login_required)
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        try:
            return super().dispatch(*args, **kwargs)
        except Exception as e:
            self._logger.error(f"Error in dispatch: {e}", exc_info=True)
            return JsonResponse({'error': 'Внутренняя ошибка сервера'}, status=500)
    
    def post(self, request):
        """Обработка POST запроса для импорта"""
        self._logger.info(f"Excel import via web by user: {request.user.username}")
        
        file = request.FILES.get('excel_file')
        
        if not file:
            return JsonResponse({'error': 'Файл не выбран'}, status=400)
        
        if not file.name.endswith(('.xlsx', '.xls')):
            self._logger.warning(f"Invalid file format: {file.name}")
            return JsonResponse(
                {'error': 'Поддерживаются только файлы Excel (.xlsx, .xls)'}, 
                status=400
            )
        
        if file.size > 10 * 1024 * 1024:
            self._logger.warning(f"File too large: {file.size} bytes")
            return JsonResponse(
                {'error': 'Размер файла не должен превышать 10 МБ'}, 
                status=413
            )
        
        organization_id = None
        if hasattr(request, 'current_organization') and request.current_organization:
            organization_id = request.current_organization.id
            self._logger.info(f"Importing to organization: {request.current_organization.short_name}")
        
        tmp_file_path = None
        
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in file.chunks():
                    tmp_file.write(chunk)
                tmp_file_path = tmp_file.name
            
            self._logger.debug(f"File saved to temp: {tmp_file_path}")
            
            importer = ExcelImporter(tmp_file_path, organization_id)
            stats = importer.import_data()
            
            self._logger.info(
                f"Web import completed: {stats.get('total_rows', 0)} rows, "
                f"{stats.get('plots_created', 0)} plots, "
                f"{stats.get('owners_created', 0)} owners, "
                f"{stats.get('errors', 0)} errors"
            )
            
            return JsonResponse({
                'success': True,
                'stats': stats,
                'errors': importer.errors[:20],
                'warnings': importer.warnings[:20],
            })
            
        except MemoryError:
            self._logger.error("Memory error during web import")
            return JsonResponse(
                {'error': 'Файл слишком большой для обработки'}, 
                status=413
            )
        except Exception as e:
            self._logger.error(f"Web import error: {e}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=500)
        finally:
            if tmp_file_path and os.path.exists(tmp_file_path):
                try:
                    os.unlink(tmp_file_path)
                    self._logger.debug(f"Temp file removed: {tmp_file_path}")
                except Exception as e:
                    self._logger.error(f"Error removing temp file: {e}")


class ExcelTemplateView(View):
    """Скачивание шаблона Excel для импорта"""
    _logger = logging.getLogger(f'{__name__}.ExcelTemplateView')
    
    def get(self, request):
        self._logger.debug("Excel template requested")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Шаблон импорта"
            
            # Заголовки
            headers = [
                '№участка', 
                'ФИО собственника', 
                'метраж', 
                'Дата выдачи', 
                'Кадастровый номер', 
                'элект.почта', 
                'телефон', 
                'примечания'
            ]
            
            # Стили для заголовков
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2c7a47', end_color='2c7a47', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Примеры данных
            examples = [
                ['1', 'Иванов Иван Иванович', 600, '15.01.2020', '50:20:0010101:123', 'ivanov@example.com', '+7(916)123-45-67', 'Участок у дороги'],
                ['2', 'Петров Петр Петрович', 800, '20.03.2019', '50:20:0010101:456', 'petrov@example.com', '+7(916)765-43-21', ''],
                ['10А', 'Сидорова Мария Петровна', 650, '10.05.2021', '', 'maria@example.com', '', 'Телефон утерян'],
                ['15', 'Кузнецов Алексей Сергеевич', 720, '01.07.2018', '50:20:0010101:789', 'kuznetsov@example.com', '+7(903)111-22-33', 'Два собственника'],
                ['', '', '', '', '', '', '', ''],
            ]
            
            # Стили для данных
            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center')
            
            # Записываем примеры
            for row_idx, row_data in enumerate(examples, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
            
            # Настраиваем ширину столбцов
            column_widths = {
                'A': 12, 'B': 30, 'C': 10, 'D': 12,
                'E': 25, 'F': 25, 'G': 20, 'H': 30,
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            ws.row_dimensions[1].height = 30
            ws.auto_filter.ref = f"A1:H{len(examples) + 1}"
            ws.freeze_panes = 'A2'
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_import.xlsx"'
            
            wb.save(response)
            
            self._logger.info("Excel template generated successfully")
            return response
            
        except Exception as e:
            self._logger.error(f"Error generating template: {e}", exc_info=True)
            return HttpResponse('Ошибка при создании шаблона', status=500)