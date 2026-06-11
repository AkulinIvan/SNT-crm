import logging
import re
from datetime import datetime
from django.db import transaction, models
from openpyxl import load_workbook
from users.models import Owner, ContactInfo
from land.models import LandPlot
from organizations.models import Organization, OrganizationMembership

logger = logging.getLogger(__name__)


class ExcelImporter:
    """Импорт данных из Excel в CRM"""
    
    def __init__(self, file_path, organization_id=None):
        self.file_path = file_path
        self.organization_id = organization_id
        self.organization = None
        self.errors = []
        self.warnings = []
        self.stats = {
            'total_rows': 0,
            'owners_created': 0,
            'owners_found': 0,
            'plots_created': 0,
            'plots_found': 0,
            'contacts_added': 0,
            'cadastral_updated': 0,
            'errors': 0
        }
    
    def normalize_cadastral(self, value):
        """
        Нормализация кадастрового номера.
        Принимает различные форматы и приводит к единому виду XX:XX:XXXXXXX:XXXX
        """
        if not value:
            return None
        
        cadastral = str(value).strip()
        
        # Если значение слишком короткое или не содержит цифр
        if len(cadastral) < 5 or not any(c.isdigit() for c in cadastral):
            return None
        
        # Заменяем точки и пробелы на двоеточия
        cadastral = cadastral.replace('.', ':').replace(' ', ':')
        
        # Убираем лишние пробелы
        cadastral = re.sub(r'\s+', '', cadastral)
        
        # Убираем лишние двоеточия
        cadastral = re.sub(r':+', ':', cadastral)
        
        # Убираем "дом", "уч" и другие слова
        cadastral = re.sub(r'дом.*$', '', cadastral, flags=re.IGNORECASE)
        cadastral = re.sub(r'уч.*$', '', cadastral, flags=re.IGNORECASE)
        
        # Если есть несколько номеров через запятую - берем первый
        if ',' in cadastral:
            cadastral = cadastral.split(',')[0].strip()
        
        # Если номер заканчивается на двоеточие - удаляем его
        cadastral = cadastral.rstrip(':')
        
        # Разбиваем на части
        parts = cadastral.split(':')
        
        # Если нет разделителей, пробуем разобрать длинный номер
        if len(parts) == 1 and ':' not in cadastral:
            digits = ''.join(c for c in cadastral if c.isdigit())
            if len(digits) >= 13:
                # Формат: 24500700439229 -> 24:50:0700439:229
                region = digits[:2]
                district = digits[2:4]
                quarter = digits[4:11]
                parcel = digits[11:]
                if quarter.startswith('0'):
                    quarter = quarter[1:]  # Убираем ведущий ноль для 7-значного квартала
                else:
                    quarter = quarter[:6]  # Берем первые 6 цифр для квартала
                return f"{region}:{district}:{quarter.zfill(6)}:{parcel.zfill(3)}"
            elif len(digits) >= 11:
                # Формат: 500700439229 -> 50:07:004392:29
                region = digits[:2]
                district = digits[2:4]
                quarter = digits[4:10] if len(digits) >= 10 else digits[4:].zfill(6)
                parcel = digits[10:] if len(digits) > 10 else '1'
                return f"{region}:{district}:{quarter.zfill(6)}:{parcel.zfill(3)}"
        
        # Обработка двухчастных номеров типа "24500700439:229"
        if len(parts) == 2:
            first_part = parts[0]
            second_part = parts[1].zfill(3)
            
            # Определяем регион (первые 2 цифры)
            region = first_part[:2]
            
            # Определяем район (следующие 2 цифры)
            if len(first_part) >= 4:
                district = first_part[2:4]
            else:
                district = '00'
            
            # Определяем квартал (оставшиеся цифры, обычно 6-7)
            remaining = first_part[4:]
            if len(remaining) >= 6:
                quarter = remaining[:7] if remaining[0] == '0' and len(remaining) >= 7 else remaining[:6]
            else:
                quarter = remaining.zfill(6)
            
            return f"{region}:{district}:{quarter.zfill(6)}:{second_part}"
        
        # Обработка трехчастных номеров
        elif len(parts) == 3:
            return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:{parts[2].zfill(6)}:001"
        
        # Обработка четырехчастных номеров
        elif len(parts) == 4:
            # Очищаем каждую часть от нецифровых символов
            cleaned_parts = []
            for part in parts:
                clean = re.sub(r'\D', '', part)
                if clean:
                    cleaned_parts.append(clean)
                else:
                    cleaned_parts.append('0')
            
            if len(cleaned_parts) == 4:
                # Форматируем каждую часть
                region = cleaned_parts[0].zfill(2)[:2]
                district = cleaned_parts[1].zfill(2)[:2]
                quarter = cleaned_parts[2].zfill(6)[:7]  # Может быть 6 или 7 цифр
                if len(quarter) == 7:
                    quarter = quarter[:6]
                parcel = cleaned_parts[3].zfill(3)[:10]
                
                return f"{region}:{district}:{quarter}:{parcel}"
        
        # Если не удалось нормализовать, возвращаем исходное значение
        return None
    
    def parse_phone(self, value):
        """Парсинг и форматирование телефона"""
        if not value:
            return None

        # Извлекаем только цифры
        digits = re.sub(r'\D', '', str(value))

        if not digits:
            return None

        # Приводим к 11-значному формату с 7 в начале
        if len(digits) == 11:
            if digits.startswith('8'):
                digits = '7' + digits[1:]  # Меняем 8 на 7
            elif digits.startswith('7'):
                pass  # Уже правильный формат
            else:
                return None
        elif len(digits) == 10:
            digits = '7' + digits  # Добавляем 7 в начало
        else:
            return None

        # Проверяем, что получился валидный номер
        if len(digits) == 11 and digits.startswith('7'):
            # Форматируем: +7 (XXX) XXX-XX-XX
            formatted = f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
            return formatted

        return None
    
    def extract_phones(self, text):
        """Извлечь все номера телефонов из текста"""
        if not text:
            return []
        
        phones = []
        text = str(text)
        
        # Сначала убираем все не-цифровые символы и ищем последовательности цифр
        # Ищем группы цифр длиной 10-11
        digit_groups = re.findall(r'\d{10,11}', re.sub(r'\D', ' ', text))
        
        for digits in digit_groups:
            phone = self.parse_phone(digits)
            if phone and phone not in phones:
                phones.append(phone)
        
        # Если не нашли, пробуем более сложные паттерны
        if not phones:
            patterns = [
                r'(\+?7[-\s]?\(?\d{3}\)?[-\s]?\d{3}[-\s]?\d{2}[-\s]?\d{2})',
                r'(8[-\s]?\(?\d{3}\)?[-\s]?\d{3}[-\s]?\d{2}[-\s]?\d{2})',
                r'(\d{3}[-\s]?\d{3}[-\s]?\d{2}[-\s]?\d{2})',
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, text)
                for match in matches:
                    phone = self.parse_phone(match)
                    if phone and phone not in phones:
                        phones.append(phone)
        
        return phones[:3]  # Возвращаем не более 3 номеров
    
    def extract_email(self, text):
        """Извлечь email из текста"""
        if not text:
            return None
        
        pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        match = re.search(pattern, str(text))
        if match:
            return match.group().lower()
        return None
    
    def parse_area(self, value):
        """Парсинг площади"""
        if not value:
            return None
        
        try:
            area_str = str(value).replace(',', '.').strip()
            # Убираем все кроме цифр и точки
            area_str = re.sub(r'[^\d.]', '', area_str)
            if area_str:
                return round(float(area_str), 2)
        except:
            pass
        return None
    
    def parse_date(self, value):
        """Парсинг даты"""
        if not value:
            return None
        
        if isinstance(value, datetime):
            return value.date()
        
        date_str = str(value).strip()
        
        # Убираем "г", "г." и другие символы
        date_str = re.sub(r'[г\.]', '', date_str, flags=re.IGNORECASE)
        date_str = re.sub(r'[^\d\.\-/]', '', date_str)
        
        formats = [
            '%d.%m.%Y', '%d.%m.%y', 
            '%Y-%m-%d', 
            '%d/%m/%Y', '%d/%m/%y',
            '%d.%m.%Y', '%d.%m.%y',
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except:
                continue
        
        return None
    
    def normalize_name(self, name):
        """Нормализация ФИО"""
        if not name:
            return None
        
        name = str(name).strip()
        
        # Убираем лишние символы
        name = re.sub(r'[?!@#$%^&*()]+', '', name)
        name = re.sub(r'\?+', '', name)
        name = re.sub(r'!+', '', name)
        
        # Убираем слова-маркеры
        markers = ['СПК', 'СТ', 'С.Т', 'СНТ', 'ПК', 'Строитель', 'Строителб']
        for marker in markers:
            name = re.sub(rf'\s+{marker}.*$', '', name, flags=re.IGNORECASE)
        
        # Убираем лишние пробелы
        name = ' '.join(name.split())
        
        # Приводим к правильному регистру (первая буква заглавная)
        name_parts = name.split()
        formatted_parts = []
        for part in name_parts:
            if part:
                formatted_parts.append(part.capitalize())
        name = ' '.join(formatted_parts)
        
        return name if len(name) > 3 else None
    
    def _get_cell_value(self, row, index, default=None):
        """Безопасное получение значения ячейки"""
        if index < len(row) and row[index] is not None:
            val = row[index]
            if isinstance(val, str):
                return val.strip()
            return val
        return default
    
    @transaction.atomic
    def import_data(self):
        """Основной метод импорта данных"""
        try:
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            # Получаем организацию
            if self.organization_id:
                try:
                    self.organization = Organization.objects.get(id=self.organization_id)
                except Organization.DoesNotExist:
                    self.errors.append(f"Организация с ID {self.organization_id} не найдена")
                    return self.stats
            
            print(f"Импорт в организацию: {self.organization.short_name if self.organization else 'Без организации'}")
            print("=" * 80)
            
            # Пропускаем заголовки (первая строка)
            start_row = 2
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not row or not any(row):
                    continue
                
                # Извлекаем данные по столбцам (индексы 0-7):
                # 0 - №участка, 1 - ФИО, 2 - метраж, 3 - дата, 
                # 4 - кадастр, 5 - email, 6 - телефон, 7 - примечания
                plot_number = self._get_cell_value(row, 0)  # A
                owner_name_raw = self._get_cell_value(row, 1)  # B
                area_raw = self._get_cell_value(row, 2)  # C
                date_raw = self._get_cell_value(row, 3)  # D
                cadastral_raw = self._get_cell_value(row, 4)  # E
                email_raw = self._get_cell_value(row, 5)  # F
                phone_raw = self._get_cell_value(row, 6)  # G
                notes_raw = self._get_cell_value(row, 7)  # H
                
                # Пропускаем строки без номера участка
                if not plot_number:
                    continue
                
                # Обработка номеров участков с буквами (18а, 44а и т.д.)
                plot_number_str = str(plot_number).strip().upper()
                
                # Нормализуем ФИО
                owner_name = self.normalize_name(owner_name_raw)
                
                # Пропускаем служебные строки
                if owner_name and ('правление' in owner_name.lower()):
                    continue
                
                self.stats['total_rows'] += 1
                print(f"\n[{row_idx}] Участок {plot_number_str} - {owner_name or 'Нет ФИО'}")
                
                try:
                    # Обработка площади
                    area = self.parse_area(area_raw)
                    if area:
                        print(f"  → Площадь: {area} м²")
                    
                    # Обработка даты
                    issue_date = self.parse_date(date_raw)
                    if issue_date:
                        print(f"  → Дата: {issue_date}")
                    
                    # Обработка кадастрового номера
                    cadastral = self.normalize_cadastral(cadastral_raw)
                    if cadastral:
                        print(f"  → Кадастровый номер: {cadastral}")
                    elif cadastral_raw:
                        print(f"  → Кадастровый номер (не удалось нормализовать): {cadastral_raw}")
                    
                    # Создаем или получаем владельца (если есть ФИО)
                    owner = None
                    if owner_name:
                        owner = self._get_or_create_owner(owner_name)
                        if owner:
                            # Обрабатываем контакты
                            self._process_contacts(owner, email_raw, phone_raw, notes_raw)
                    
                    # Создаем или получаем участок
                    plot = self._get_or_create_plot(plot_number_str, cadastral, area, notes_raw)
                    
                    if plot:
                        # Обновляем кадастровый номер если он есть и отличается
                        if cadastral and (not plot.cadastral_number or plot.cadastral_number != cadastral):
                            # Проверяем, не занят ли кадастровый номер другим участком
                            existing = LandPlot.objects.filter(cadastral_number=cadastral).exclude(id=plot.id).first()
                            if existing:
                                self._add_warning(row_idx, f"Кадастровый номер {cadastral} уже используется участком {existing.plot_number}")
                            else:
                                plot.cadastral_number = cadastral
                                plot.save(update_fields=['cadastral_number'])
                                self.stats['cadastral_updated'] += 1
                                print(f"  → Кадастровый номер обновлен: {cadastral}")
                        
                        # Обновляем площадь если она есть и отличается
                        if area and (not plot.area_sqm or abs(plot.area_sqm - area) > 0.01):
                            plot.area_sqm = area
                            plot.save(update_fields=['area_sqm'])
                            print(f"  → Площадь обновлена: {area} м²")
                        
                        # Связываем участок с владельцем
                        if owner:
                            from users.models import Ownership
                            ownership_exists = Ownership.objects.filter(
                                owner=owner, 
                                land_plot=plot
                            ).exists()
                            
                            if not ownership_exists:
                                Ownership.objects.create(
                                    owner=owner,
                                    land_plot=plot,
                                    share='1/1',
                                    ownership_since=issue_date,
                                    document_basis=f"Импорт из Excel {datetime.now().strftime('%d.%m.%Y')}"
                                )
                                print(f"  → Связан с владельцем")
                            else:
                                print(f"  → Связь с владельцем уже существует")
                    
                except Exception as e:
                    self._add_error(row_idx, f"Ошибка: {str(e)}")
                    logger.exception(f"Ошибка в строке {row_idx}")
            
            print("\n" + "=" * 80)
            print("РЕЗУЛЬТАТЫ ИМПОРТА:")
            print(f"  Всего строк: {self.stats['total_rows']}")
            print(f"  Создано владельцев: {self.stats['owners_created']}")
            print(f"  Найдено владельцев: {self.stats['owners_found']}")
            print(f"  Создано участков: {self.stats['plots_created']}")
            print(f"  Найдено участков: {self.stats['plots_found']}")
            print(f"  Добавлено контактов: {self.stats['contacts_added']}")
            print(f"  Обновлено кадастровых номеров: {self.stats['cadastral_updated']}")
            print(f"  Ошибок: {self.stats['errors']}")
            print(f"  Предупреждений: {len(self.warnings)}")
            
            return self.stats
            
        except Exception as e:
            self.errors.append(f"Ошибка открытия файла: {str(e)}")
            logger.exception("Ошибка импорта")
            return self.stats
    
    def _process_contacts(self, owner, email_raw, phone_raw, notes_raw):
        """Обработка контактов"""
        # Обработка email из отдельного столбца
        if email_raw:
            email = self.extract_email(email_raw)
            if email:
                self._add_contact(owner, 'em', email)
                print(f"  → Email: {email}")
        
        # Обработка телефона из отдельного столбца
        if phone_raw:
            phones = self.extract_phones(phone_raw)
            for phone in phones[:2]:  # Максимум 2 телефона
                if phone:
                    self._add_contact(owner, 'ph', phone)
                    print(f"  → Телефон: {phone}")
        
        # Поиск дополнительных контактов в примечаниях
        if notes_raw:
            # Ищем телефоны в примечаниях
            phones = self.extract_phones(notes_raw)
            for phone in phones[:3]:
                if phone:
                    self._add_contact(owner, 'ph', phone)
                    print(f"  → Телефон из примечаний: {phone}")
            
            # Ищем email в примечаниях
            if not email_raw:
                email = self.extract_email(notes_raw)
                if email:
                    self._add_contact(owner, 'em', email)
                    print(f"  → Email из примечаний: {email}")
    
    def _get_or_create_owner(self, full_name):
        """Получить или создать владельца"""
        # Очищаем имя для поиска
        search_name = full_name.replace('ё', 'е')
        
        # Ищем существующего владельца (точное совпадение)
        owner = Owner.objects.filter(full_name__iexact=full_name).first()
        
        if owner:
            self.stats['owners_found'] += 1
            return owner
        
        # Ищем похожего владельца
        owner = Owner.objects.filter(full_name__icontains=full_name[:20]).first()
        
        if owner:
            self.stats['owners_found'] += 1
            self._add_warning(0, f"Найден похожий владелец для '{full_name}': {owner.full_name}")
            return owner
        
        # Создаем нового владельца
        try:
            owner = Owner.objects.create(full_name=full_name)
            self.stats['owners_created'] += 1
            print(f"  ✓ Создан владелец: {full_name}")
            
            # Если есть организация, создаем членство
            if self.organization:
                OrganizationMembership.objects.get_or_create(
                    owner=owner,
                    organization=self.organization,
                    defaults={'status': 'active'}
                )
            
            return owner
        except Exception as e:
            self._add_error(0, f"Ошибка создания владельца '{full_name}': {str(e)}")
            return None
    
    def _get_or_create_plot(self, plot_number, cadastral_number, area_sqm, notes):
        """Получить или создать участок"""
        # Проверка лимитов перед созданием
        if self.organization:
            is_allowed, current, max_limit, message = self.organization.check_tariff_limit('plots')
            if not is_allowed:
                self._add_error(0, f"Лимит участков ({current}/{max_limit}) превышен. {message}")
                return None
        # Очищаем номер участка
        plot_number = str(plot_number).strip().upper()
        
        # Ищем по номеру участка
        plot = LandPlot.objects.filter(plot_number=plot_number).first()
        
        if plot:
            self.stats['plots_found'] += 1
            return plot
        
        # Если нет кадастрового номера, создаем временный
        if not cadastral_number:
            # Генерируем уникальный временный кадастровый номер
            import time
            timestamp = str(int(time.time()))[-6:]
            cadastral_number = f"00:00:000000:{timestamp}"
            self._add_warning(0, f"Временный кадастровый номер для участка {plot_number}: {cadastral_number}")
        
        # Проверяем уникальность кадастрового номера
        existing = LandPlot.objects.filter(cadastral_number=cadastral_number).first()
        if existing:
            self._add_error(0, f"Кадастровый номер {cadastral_number} уже занят участком {existing.plot_number}")
            # Создаем с другим кадастровым номером
            import time
            timestamp = str(int(time.time()))[-6:]
            cadastral_number = f"00:00:000000:{timestamp}"
        
        # Создаем новый участок
        try:
            plot = LandPlot.objects.create(
                plot_number=plot_number,
                cadastral_number=cadastral_number,
                area_sqm=area_sqm or 600.0,
                notes=f"Импортировано из Excel. {notes[:200] if notes else ''}"[:500],
                status='active',
                organization=self.organization
            )
            self.stats['plots_created'] += 1
            print(f"  ✓ Создан участок: {plot_number} (площадь: {area_sqm or 600} м²)")
            return plot
        except Exception as e:
            self._add_error(0, f"Ошибка создания участка {plot_number}: {str(e)}")
            return None
    
    def _add_contact(self, owner, contact_type, value):
        """Добавить контакт владельцу"""
        if not value:
            return
        
        # Проверяем, существует ли уже такой активный контакт
        existing = ContactInfo.objects.filter(
            owner=owner,
            type=contact_type,
            value=value,
            is_active=True
        ).first()
        
        if existing:
            return
        
        # Создаем новый контакт
        try:
            ContactInfo.objects.create(
                owner=owner,
                type=contact_type,
                value=value,
                is_active=True,
                is_verified=False,
                note="Импортировано из Excel"
            )
            self.stats['contacts_added'] += 1
        except Exception as e:
            self._add_warning(0, f"Не удалось добавить контакт {value}: {str(e)}")
    
    def _add_error(self, row, message):
        """Добавить ошибку"""
        error_msg = f"Строка {row}: {message}" if row > 0 else message
        self.errors.append(error_msg)
        self.stats['errors'] += 1
    
    def _add_warning(self, row, message):
        """Добавить предупреждение"""
        warning_msg = f"Строка {row}: {message}" if row > 0 else message
        self.warnings.append(warning_msg)