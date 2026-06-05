import re
import csv
import json
import logging
import os
import traceback
from datetime import date, datetime
from typing import List, Dict, Optional, Tuple, Any
from decimal import Decimal
from pathlib import Path

# Настройка логгера с записью в файл
logger = logging.getLogger('payments.bank_parser')

# Настройка файлового логгера для детального логирования
file_handler = logging.FileHandler('logs/bank_parser.log', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s'
)
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

# Консольный логгер
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)


class BankStatementParser:
    """
    Универсальный парсер банковских выписок.
    Поддерживает форматы: CSV, Excel, 1С, JSON, PDF.
    """
    
    # Паттерны для определения банка по формату файла
    BANK_PATTERNS = {
        'sberbank': {
            'date_col': ['Дата операции', 'Дата', 'date'],
            'amount_col': ['Сумма', 'Сумма операции', 'amount'],
            'payer_col': ['Плательщик', 'Контрагент', 'payer', 'Наименование'],
            'account_col': ['Счет плательщика', 'Счет', 'account'],
            'inn_col': ['ИНН плательщика', 'ИНН', 'inn'],
            'purpose_col': ['Назначение платежа', 'Назначение', 'purpose'],
        },
        'tinkoff': {
            'date_col': ['Дата операции', 'date'],
            'amount_col': ['Сумма', 'amount'],
            'payer_col': ['Отправитель', 'Контрагент', 'name'],
            'account_col': ['Счет отправителя', 'account'],
            'inn_col': ['ИНН', 'inn'],
            'purpose_col': ['Назначение', 'purpose'],
        },
        'alfa': {
            'date_col': ['Дата', 'date'],
            'amount_col': ['Сумма в валюте счета', 'amount'],
            'payer_col': ['Наименование получателя/плательщика', 'name'],
            'account_col': ['Счет', 'account'],
            'inn_col': ['ИНН', 'inn'],
            'purpose_col': ['Назначение платежа', 'purpose'],
        },
    }
    
    def __init__(self, bank_name: Optional[str] = None):
        self.bank_name = bank_name
        self.bank_pattern = self.BANK_PATTERNS.get(bank_name, {})
        logger.info(f"BankStatementParser initialized with bank_name={bank_name}")
    
    def detect_bank(self, file_path: str) -> str:
        """Автоматически определить банк по структуре файла"""
        logger.info(f"Detecting bank for file: {file_path}")
        
        try:
            if not os.path.exists(file_path):
                logger.error(f"File not found: {file_path}")
                return 'unknown'
            
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = [h.lower().strip() for h in next(reader)]
                    logger.debug(f"CSV headers: {headers}")
                    
                    for bank, patterns in self.BANK_PATTERNS.items():
                        matches = 0
                        for pattern_list in patterns.values():
                            if any(p.lower() in headers for p in pattern_list):
                                matches += 1
                        if matches >= 3:
                            logger.info(f"Bank detected: {bank}")
                            return bank
            
            logger.info("Bank not detected, returning 'unknown'")
            return 'unknown'
            
        except Exception as e:
            logger.error(f"Error detecting bank: {e}\n{traceback.format_exc()}")
            return 'unknown'
    
    def parse_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Парсинг файла выписки с улучшенным логированием"""
        logger.info(f"Starting to parse file: {file_path}")
        
        try:
            # Проверка существования файла
            if not file_path:
                logger.error("File path is empty")
                return []
            
            if not os.path.exists(file_path):
                logger.error(f"File does not exist: {file_path}")
                return []
            
            # Проверка размера файла
            file_size = os.path.getsize(file_path)
            logger.info(f"File size: {file_size} bytes")
            
            if file_size == 0:
                logger.error("File is empty")
                return []
            
            # Определяем тип файла
            file_ext = os.path.splitext(file_path)[1].lower()
            logger.info(f"File extension: {file_ext}")
            
            if file_ext == '.pdf':
                logger.info("Detected PDF file format")
                transactions = self._parse_pdf(file_path)
                if transactions is None:
                    transactions = []
                logger.info(f"PDF parsing completed, found {len(transactions)} transactions")

                # Выводим каждую транзакцию для отладки
                for i, trans in enumerate(transactions):
                    logger.debug(f"Transaction {i+1}: date={trans.get('transaction_date')}, "
                               f"amount={trans.get('amount')}, payer={trans.get('payer_name')}, "
                               f"UID={trans.get('matched_uid')}")

                return transactions

            elif file_ext == '.csv':
                logger.info("Detected CSV file format")
                return self._parse_csv(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                logger.info("Detected Excel file format")
                return self._parse_excel(file_path)
            elif file_ext == '.json':
                logger.info("Detected JSON file format")
                return self._parse_json(file_path)
            elif file_ext == '.txt':
                logger.info("Detected TXT file format (1C format)")
                return self._parse_1c(file_path)
            else:
                logger.warning(f"Unsupported file format: {file_ext}")
                raise ValueError(f'Неподдерживаемый формат файла: {file_path}')
                
        except Exception as e:
            logger.error(f"Parse error: {e}\n{traceback.format_exc()}")
            return []  # Всегда возвращаем список, даже при ошибке
    
    def _parse_alfa_receipt(self, text: str) -> List[Dict[str, Any]]:
        """
        Специальный парсер для квитанций Альфа-Банка
        """
        transactions = []
        logger.info("=" * 50)
        logger.info("Parsing Alfa-Bank receipt...")
        
        try:
            # Нормализуем текст - разбиваем на строки
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            logger.debug(f"Total lines in receipt: {len(lines)}")

            # Извлекаем сумму
            amount = None
            for line in lines:
                # Ищем строку с суммой перевода
                if 'Сумма перевода' in line or 'Сумма' in line:
                    # Ищем число с пробелами и запятой
                    match = re.search(r'(\d[\d\s]*,\d{2})\s*RUR', line)
                    if match:
                        amount_str = match.group(1).replace(' ', '').replace(',', '.')
                        amount = Decimal(amount_str)
                        logger.info(f"Found amount: {amount}")
                        break
                        
            if amount is None:
                logger.error("Amount not found in receipt")
                return []

            # Извлекаем дату
            transaction_date = date.today()
            for line in lines:
                if 'Дата и время перевода' in line:
                    # Ищем дату на следующей строке или в этой же
                    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', line)
                    if date_match:
                        date_str = date_match.group(1)
                        try:
                            transaction_date = datetime.strptime(date_str, '%d.%m.%Y').date()
                            logger.info(f"Found date: {transaction_date}")
                            break
                        except ValueError as e:
                            logger.warning(f"Error parsing date: {e}")
                            
            # Извлекаем плательщика (улучшенный поиск)
            payer_name = ''

            # Ищем строку "Плательщик" и берем следующую
            for i, line in enumerate(lines):
                if line == 'Плательщик' or line.startswith('Плательщик'):
                    # Следующая строка должна содержать ФИО
                    if i + 1 < len(lines):
                        potential_name = lines[i + 1]
                        # Проверяем, что это похоже на ФИО (содержит буквы и пробелы)
                        if re.match(r'^[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+$', potential_name):
                            payer_name = potential_name
                            logger.info(f"Found payer (from 'Плательщик'): {payer_name}")
                            break
                        elif re.match(r'^[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+$', potential_name):
                            payer_name = potential_name
                            logger.info(f"Found payer (from 'Плательщик' - short): {payer_name}")
                            break
                        
            # Если не нашли, ищем ФИО в тексте
            if not payer_name:
                # Ищем ФИО в формате "Фамилия Имя Отчество"
                name_pattern = r'([А-ЯЁ][а-яё]+)\s+([А-ЯЁ][а-яё]+)\s+([А-ЯЁ][а-яё]+)'
                matches = re.findall(name_pattern, text)
                for match in matches:
                    full_name = f"{match[0]} {match[1]} {match[2]}"
                    # Исключаем "Строитель-43" и другие не-ФИО
                    if 'Строитель' not in full_name and len(full_name) > 10:
                        payer_name = full_name
                        logger.info(f"Found payer (by pattern): {payer_name}")
                        break
                    
            # Извлекаем назначение платежа
            payment_purpose = ''
            for i, line in enumerate(lines):
                if 'Назначение перевода' in line:
                    # Собираем следующие строки до следующего поля
                    purpose_lines = []
                    for j in range(i + 1, len(lines)):
                        if 'Плательщик' in lines[j] or 'Получатель' in lines[j] or 'Банк' in lines[j]:
                            break
                        purpose_lines.append(lines[j])
                    payment_purpose = ' '.join(purpose_lines)
                    payment_purpose = ' '.join(payment_purpose.split())
                    logger.info(f"Found payment purpose: {payment_purpose[:100]}...")
                    break
                
            # Если не нашли назначение, ищем в остатке текста
            if not payment_purpose:
                purpose_match = re.search(r'Назначение\s+перевода\s+(.+?)(?=\s+[А-Я][а-я]+:|$)', text, re.DOTALL)
                if purpose_match:
                    payment_purpose = purpose_match.group(1).strip()
                    payment_purpose = ' '.join(payment_purpose.split())
                    logger.info(f"Found payment purpose (fallback): {payment_purpose[:100]}...")

            # Извлекаем номер операции
            transaction_id = ''
            for line in lines:
                if 'Номер операции' in line:
                    match = re.search(r'(\w+)', line.replace('Номер операции', ''))
                    if match:
                        transaction_id = match.group(1)
                        logger.info(f"Found transaction ID: {transaction_id}")
                        break
                    
            # Извлекаем UID начисления
            matched_uid = ''
            snt_match = re.search(r'SNT-\d{6}', payment_purpose)
            if snt_match:
                matched_uid = snt_match.group(0)
                logger.info(f"Found UID in purpose: {matched_uid}")

            # Если не нашли в назначении, ищем во всём тексте
            if not matched_uid:
                snt_match = re.search(r'ID:?(SNT-\d{6})', text)
                if snt_match:
                    matched_uid = snt_match.group(1)
                    logger.info(f"Found UID in text: {matched_uid}")

            # Извлекаем номер участка
            plot_number = ''
            plot_match = re.search(r'Уч\.№(\d+)', payment_purpose)
            if plot_match:
                plot_number = plot_match.group(1)
                logger.debug(f"Found plot number: {plot_number}")

            transaction = {
                'transaction_date': transaction_date,
                'amount': amount,
                'payer_name': payer_name,
                'payment_purpose': payment_purpose,
                'transaction_id': transaction_id,
                'matched_uid': matched_uid,
                'plot_number': plot_number,
            }

            logger.info(f"✅ Successfully parsed payment: {amount} ₽ from {payer_name or 'unknown'}, UID: {matched_uid}")
            transactions.append(transaction)
            
        except Exception as e:
            logger.error(f"Error parsing Alfa receipt: {e}\n{traceback.format_exc()}")
            
        return transactions

    
    def _parse_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Парсинг PDF выписки или квитанции
        """
        logger.info(f"Starting PDF parsing: {file_path}")
        
        try:
            # Сначала пробуем извлечь таблицы
            transactions = self._parse_pdf_table(file_path)
            if transactions and len(transactions) > 0:
                logger.info(f'Extracted {len(transactions)} transactions from PDF tables')
                return transactions

            # Если таблиц нет - извлекаем текст
            text = self._extract_text_from_pdf(file_path)

            # Проверяем, что текст не пустой
            if not text or not text.strip():
                logger.error("Failed to extract text from PDF - file is empty or protected")
                return []

            logger.info(f'Extracted text from PDF ({len(text)} characters)')
            logger.debug(f'First 500 characters: {text[:500]}')

            # Пробуем разные форматы

            # 1. Квитанция Альфа-Банка
            if 'Квитанция о переводе' in text or 'АО "АЛЬФА-БАНК"' in text:
                logger.info("Detected Alfa-Bank receipt format")
                result = self._parse_alfa_receipt(text)
                if result:
                    return result
                else:
                    logger.warning("Failed to parse as Alfa-Bank receipt, trying other formats")

            # 2. Квитанция Сбербанка
            if 'Сбербанк' in text and 'Квитанция' in text:
                logger.info("Detected Sberbank receipt format")
                result = self._parse_sberbank_receipt(text)
                if result:
                    return result

            # 3. Выписка Сбербанка
            text_lower = text.lower()
            if 'сбербанк' in text_lower or 'sberbank' in text_lower:
                logger.info("Detected Sberbank statement format")
                return self._parse_sberbank_text(text)
            elif 'тинькофф' in text_lower or 'tinkoff' in text_lower:
                logger.info("Detected Tinkoff statement format")
                return self._parse_tinkoff_text(text)
            elif 'альфа' in text_lower or 'alfa' in text_lower:
                logger.info("Detected Alfa-Bank statement format")
                return self._parse_alfa_text(text)
            else:
                logger.info("Using generic text parser")
                return self._parse_generic_text(text)
                
        except Exception as e:
            logger.error(f"Error parsing PDF: {e}\n{traceback.format_exc()}")
            return []
    
    def _parse_sberbank_receipt(self, text: str) -> List[Dict[str, Any]]:
        """
        Парсер для квитанций Сбербанка
        """
        transactions = []
        logger.info("Parsing Sberbank receipt...")
        
        try:
            # Извлекаем сумму
            amount_patterns = [
                r'Сумма\s+перевода[:\s]*([\d\s]+,\d{2})\s*₽',
                r'Сумма[:\s]*([\d\s]+,\d{2})\s*₽',
            ]
            
            amount = None
            for pattern in amount_patterns:
                amount_match = re.search(pattern, text)
                if amount_match:
                    amount_str = amount_match.group(1).replace(' ', '').replace(',', '.')
                    amount = Decimal(amount_str)
                    logger.info(f"Found amount: {amount}")
                    break
            
            if amount is None:
                logger.error("Amount not found in Sberbank receipt")
                return []
            
            # Извлекаем дату
            date_pattern = r'Дата\s+(\d{2}\.\d{2}\.\d{4})'
            date_match = re.search(date_pattern, text)
            if date_match:
                date_str = date_match.group(1)
                transaction_date = datetime.strptime(date_str, '%d.%m.%Y').date()
                logger.info(f"Found date: {transaction_date}")
            else:
                transaction_date = date.today()
                logger.warning(f"Date not found, using today: {transaction_date}")
            
            # Извлекаем плательщика
            payer_pattern = r'Плательщик[:\s]+([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)'
            payer_match = re.search(payer_pattern, text)
            payer_name = payer_match.group(1) if payer_match else ''
            logger.info(f"Found payer: {payer_name}")
            
            # Извлекаем назначение
            purpose_pattern = r'Назначение\s+платежа[:\s]+(.+?)(?:\n|$)'
            purpose_match = re.search(purpose_pattern, text, re.DOTALL)
            payment_purpose = purpose_match.group(1).strip() if purpose_match else ''
            logger.debug(f"Found purpose: {payment_purpose[:100]}...")
            
            # Номер операции
            operation_pattern = r'Номер\s+операции[:\s]+(\w+)'
            operation_match = re.search(operation_pattern, text)
            transaction_id = operation_match.group(1) if operation_match else ''
            
            # UID начисления
            snt_id_pattern = r'SNT-(\d{6})'
            snt_match = re.search(snt_id_pattern, payment_purpose)
            matched_uid = f"SNT-{snt_match.group(1)}" if snt_match else ''
            if matched_uid:
                logger.info(f"Found UID: {matched_uid}")
            
            transaction = {
                'transaction_date': transaction_date,
                'amount': amount,
                'payer_name': payer_name,
                'payment_purpose': payment_purpose,
                'transaction_id': transaction_id,
                'matched_uid': matched_uid,
            }
            
            transactions.append(transaction)
            logger.info(f"Successfully parsed Sberbank receipt: {amount} ₽")
            
        except Exception as e:
            logger.error(f"Error parsing Sberbank receipt: {e}\n{traceback.format_exc()}")
            
        return transactions

    def _extract_text_from_pdf(self, file_path: str) -> str:
        """
        Извлечение текста из PDF с использованием нескольких методов
        Всегда возвращает строку (пустую в случае ошибки)
        """
        logger.info(f"Extracting text from PDF: {file_path}")
        text = ""

        # Метод 1: pdfplumber
        try:
            import pdfplumber
            logger.info("Trying pdfplumber...")
            with pdfplumber.open(file_path) as pdf:
                logger.info(f"PDF has {len(pdf.pages)} pages")
                for page_num, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                        logger.info(f"Page {page_num}: extracted {len(page_text)} characters")
                    else:
                        # Пробуем extract_text с другими параметрами
                        page_text = page.extract_text(layout=True)
                        if page_text:
                            text += page_text + "\n"
                            logger.info(f"Page {page_num} (layout): extracted {len(page_text)} characters")

            if text and text.strip():
                logger.info(f"pdfplumber extracted {len(text)} characters")
                return text
            else:
                logger.warning("pdfplumber extracted no text")
                
        except ImportError:
            logger.warning('pdfplumber not installed')
        except Exception as e:
            logger.error(f'pdfplumber error: {e}')

        # Метод 2: PyPDF2
        try:
            from PyPDF2 import PdfReader
            logger.info("Trying PyPDF2...")
            reader = PdfReader(file_path)
            logger.info(f"PDF has {len(reader.pages)} pages")
            for page_num, page in enumerate(reader.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                    logger.info(f"Page {page_num}: {len(page_text)} characters")

            if text and text.strip():
                logger.info(f"PyPDF2 extracted {len(text)} characters")
                return text
            else:
                logger.warning("PyPDF2 extracted no text")
                
        except ImportError:
            logger.warning('PyPDF2 not installed')
        except Exception as e:
            logger.error(f'PyPDF2 error: {e}')

        # Метод 3: pdfminer
        try:
            from pdfminer.high_level import extract_text
            logger.info("Trying pdfminer...")
            extracted = extract_text(file_path)
            if extracted and extracted.strip():
                text = extracted
                logger.info(f"pdfminer extracted {len(text)} characters")
                return text
            else:
                logger.warning("pdfminer extracted no text")
                
        except ImportError:
            logger.warning('pdfminer not installed')
        except Exception as e:
            logger.error(f'pdfminer error: {e}')

        # Метод 4: пробуем прочитать как бинарный файл
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read()
                # Пробуем декодировать как utf-8
                try:
                    decoded = raw_data.decode('utf-8', errors='ignore')
                    if decoded and decoded.strip():
                        logger.info(f"RAW text extracted: {len(decoded)} characters")
                        return decoded
                except Exception as e:
                    logger.warning(f"Failed to decode as UTF-8: {e}")
        except Exception as e:
            logger.error(f'Error reading RAW: {e}')

        # Всегда возвращаем строку, даже пустую
        logger.warning("Failed to extract text from PDF, returning empty string")
        return ""  # Никогда не возвращаем None
    
    def _parse_sberbank_text(self, text: str) -> List[Dict[str, Any]]:
        """
        Парсинг текста выписки Сбербанка.
        """
        transactions = []
        logger.info("Parsing Sberbank text statement...")
        
        try:
            lines = text.split('\n')
            logger.debug(f"Total lines: {len(lines)}")
            
            # Ищем строки с датами и суммами
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue
                
                # Паттерн: дата + сумма + текст
                date_pattern = r'(\d{2}\.\d{2}\.\d{4})'
                amount_pattern = r'([\d\s]+[.,]\d{2})'
                
                date_match = re.search(date_pattern, line)
                amount_match = re.search(amount_pattern, line)
                
                if date_match and amount_match:
                    date_str = date_match.group(1)
                    amount_str = amount_match.group(1).replace(' ', '').replace(',', '.')
                    
                    # Всё после даты и суммы — плательщик и назначение
                    rest_of_line = line[amount_match.end():].strip()
                    
                    # Разделяем по разделителям Сбербанка
                    parts = re.split(r'\s{2,}|\t|\|', rest_of_line)
                    
                    payer_name = parts[0] if parts else ''
                    payment_purpose = ' '.join(parts[1:]) if len(parts) > 1 else ''
                    
                    try:
                        amount = Decimal(amount_str)
                        if amount > 0:  # Только поступления
                            transactions.append({
                                'transaction_date': datetime.strptime(date_str, '%d.%m.%Y').date(),
                                'amount': amount,
                                'payer_name': payer_name.strip(),
                                'payment_purpose': payment_purpose.strip(),
                            })
                            logger.debug(f"Found transaction at line {line_num}: {amount} ₽ from {payer_name[:30]}")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Error parsing amount at line {line_num}: {e}")
                        continue
            
            logger.info(f'Found {len(transactions)} transactions in Sberbank statement')
            
        except Exception as e:
            logger.error(f"Error parsing Sberbank text: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _parse_tinkoff_text(self, text: str) -> List[Dict[str, Any]]:
        """Парсинг текста выписки Тинькофф"""
        transactions = []
        logger.info("Parsing Tinkoff text statement...")
        
        try:
            lines = text.split('\n')
            
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue
                
                date_pattern = r'(\d{2}\.\d{2}\.\d{4})'
                amount_pattern = r'([\d\s]+[.,]\d{2})\s*(?:₽|руб|RUB)?'
                
                date_match = re.search(date_pattern, line)
                amount_match = re.search(amount_pattern, line)
                
                if date_match and amount_match:
                    try:
                        date_str = date_match.group(1)
                        amount_str = amount_match.group(1).replace(' ', '').replace(',', '.')
                        amount = Decimal(amount_str)
                        
                        if amount > 0:
                            # Ищем плательщика (обычно после суммы)
                            rest = line[amount_match.end():].strip()
                            
                            transactions.append({
                                'transaction_date': datetime.strptime(date_str, '%d.%m.%Y').date(),
                                'amount': amount,
                                'payer_name': rest[:100] if rest else '',
                                'payment_purpose': rest if rest else '',
                            })
                            logger.debug(f"Found transaction at line {line_num}: {amount} ₽")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Error parsing at line {line_num}: {e}")
                        continue
            
            logger.info(f'Found {len(transactions)} transactions in Tinkoff statement')
            
        except Exception as e:
            logger.error(f"Error parsing Tinkoff text: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _parse_alfa_text(self, text: str) -> List[Dict[str, Any]]:
        """Парсинг текста выписки Альфа-Банк"""
        transactions = []
        logger.info("Parsing Alfa-Bank text statement...")
        
        try:
            lines = text.split('\n')
            
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue
                
                date_pattern = r'(\d{2}\.\d{2}\.\d{4})'
                amount_pattern = r'([\d\s]+[.,]\d{2})'
                
                date_match = re.search(date_pattern, line)
                amount_match = re.search(amount_pattern, line)
                
                if date_match and amount_match:
                    try:
                        date_str = date_match.group(1)
                        amount_str = amount_match.group(1).replace(' ', '').replace(',', '.')
                        amount = Decimal(amount_str)
                        
                        if amount > 0:
                            rest = line[amount_match.end():].strip()
                            transactions.append({
                                'transaction_date': datetime.strptime(date_str, '%d.%m.%Y').date(),
                                'amount': amount,
                                'payer_name': rest[:100] if rest else '',
                                'payment_purpose': rest if rest else '',
                            })
                            logger.debug(f"Found transaction at line {line_num}: {amount} ₽")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Error parsing at line {line_num}: {e}")
                        continue
            
            logger.info(f'Found {len(transactions)} transactions in Alfa-Bank statement')
            
        except Exception as e:
            logger.error(f"Error parsing Alfa-Bank text: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _parse_generic_text(self, text: str) -> List[Dict[str, Any]]:
        """
        Универсальный парсинг текста выписки.
        """
        transactions = []
        logger.info("Using generic text parser...")
        
        try:
            lines = text.split('\n')
            
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue
                
                # Ищем дату
                date_pattern = r'(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})'
                amount_pattern = r'([\d\s]+[.,]\d{2})'
                
                date_match = re.search(date_pattern, line)
                amount_match = re.search(amount_pattern, line)
                
                if date_match and amount_match:
                    try:
                        date_str = date_match.group(1)
                        amount_str = amount_match.group(1).replace(' ', '').replace(',', '.')
                        amount = Decimal(amount_str)
                        
                        if amount > 0:  # Только поступления
                            # Извлекаем текст между датой и суммой (плательщик)
                            # и после суммы (назначение)
                            parts = re.split(r'\s{2,}|\t|\|', line)
                            
                            payer_name = ''
                            payment_purpose = ''
                            
                            if len(parts) >= 2:
                                # Первая часть с датой
                                # Ищем часть с именем (обычно после даты)
                                for part in parts:
                                    if re.search(r'[а-яА-ЯёЁ]', part) and not re.search(r'^\d', part):
                                        payer_name = part.strip()
                                        break
                                
                                # Назначение — всё остальное
                                other_parts = [p for p in parts if p != payer_name and not re.search(date_pattern, p) and not re.search(amount_pattern.replace('([', '(['), p)]
                                payment_purpose = ' '.join(other_parts).strip()
                            
                            transactions.append({
                                'transaction_date': datetime.strptime(date_str, '%d.%m.%Y').date(),
                                'amount': amount,
                                'payer_name': payer_name,
                                'payment_purpose': payment_purpose,
                            })
                            logger.debug(f"Found transaction at line {line_num}: {amount} ₽")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Error parsing at line {line_num}: {e}")
                        continue
            
            logger.info(f'Generic parser found {len(transactions)} transactions')
            
        except Exception as e:
            logger.error(f"Error in generic parser: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _parse_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Парсинг CSV выписки"""
        transactions = []
        logger.info(f"Parsing CSV file: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                headers = {h.lower().strip(): h for h in reader.fieldnames if h}
                logger.debug(f"CSV headers: {list(headers.keys())}")
                
                row_count = 0
                for row in reader:
                    row_count += 1
                    transaction = self._extract_transaction(row, headers)
                    if transaction:
                        transactions.append(transaction)
                        logger.debug(f"Extracted transaction {row_count}: {transaction.get('amount')} ₽")
                
                logger.info(f"Processed {row_count} rows, extracted {len(transactions)} transactions")
                
        except Exception as e:
            logger.error(f"Error parsing CSV: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Парсинг Excel выписки"""
        transactions = []
        logger.info(f"Parsing Excel file: {file_path}")
        
        try:
            import openpyxl
        except ImportError:
            logger.error('openpyxl not installed. Please run: pip install openpyxl')
            raise ImportError('Установите openpyxl: pip install openpyxl')
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            logger.info(f"Excel has {ws.max_row} rows and {ws.max_column} columns")
            
            headers = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(1, col).value
                if header:
                    headers[header.lower().strip()] = header
            
            logger.debug(f"Excel headers: {list(headers.keys())}")
            
            row_count = 0
            for row in range(2, ws.max_row + 1):
                row_count += 1
                row_data = {}
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(1, col).value
                    value = ws.cell(row, col).value
                    if header:
                        row_data[header] = value
                
                transaction = self._extract_transaction(row_data, headers)
                if transaction:
                    transactions.append(transaction)
                    logger.debug(f"Extracted transaction {row_count}: {transaction.get('amount')} ₽")
            
            logger.info(f"Processed {row_count} rows, extracted {len(transactions)} transactions")
            
        except Exception as e:
            logger.error(f"Error parsing Excel: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _parse_json(self, file_path: str) -> List[Dict[str, Any]]:
        """Парсинг JSON выписки"""
        transactions = []
        logger.info(f"Parsing JSON file: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            operations = data.get('operations', data.get('transactions', []))
            logger.info(f"Found {len(operations)} operations in JSON")
            
            for op in operations:
                try:
                    transaction = {
                        'transaction_date': self._parse_date(op.get('date', op.get('operationTime'))),
                        'amount': Decimal(str(op.get('amount', 0))),
                        'payer_name': op.get('description', op.get('counterparty', '')),
                        'payer_account': op.get('account', op.get('payerAccount', '')),
                        'payer_inn': op.get('inn', op.get('payerInn', '')),
                        'payment_purpose': op.get('purpose', op.get('paymentPurpose', '')),
                    }
                    transactions.append(transaction)
                except Exception as e:
                    logger.warning(f"Error processing JSON operation: {e}")
                    continue
            
            logger.info(f"Extracted {len(transactions)} transactions from JSON")
            
        except Exception as e:
            logger.error(f"Error parsing JSON: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _parse_1c(self, file_path: str) -> List[Dict[str, Any]]:
        """Парсинг выписки 1С (текстовый формат)"""
        transactions = []
        logger.info(f"Parsing 1C file: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            pattern = r'(\d{2}\.\d{2}\.\d{4}).*?(\d+[\.,]\d{2}).*?Плательщик:\s*(.*?)(?:\n|$)'
            matches = re.findall(pattern, content, re.MULTILINE)
            logger.info(f"Found {len(matches)} matches in 1C file")
            
            for match in matches:
                date_str, amount_str, payer = match
                try:
                    transaction = {
                        'transaction_date': datetime.strptime(date_str, '%d.%m.%Y').date(),
                        'amount': Decimal(amount_str.replace(',', '.')),
                        'payer_name': payer.strip(),
                        'payment_purpose': '',
                    }
                    transactions.append(transaction)
                except Exception as e:
                    logger.warning(f"Error processing 1C record: {e}")
                    continue
            
            logger.info(f"Extracted {len(transactions)} transactions from 1C file")
            
        except Exception as e:
            logger.error(f"Error parsing 1C file: {e}\n{traceback.format_exc()}")
            
        return transactions
    
    def _extract_transaction(self, row: Dict[str, Any], headers: Dict[str, str]) -> Optional[Dict[str, Any]]:
        """Извлечь данные транзакции из строки"""
        try:
            date_val = self._get_field(row, headers, 'date_col')
            amount_val = self._get_field(row, headers, 'amount_col')
            payer_val = self._get_field(row, headers, 'payer_col')
            
            if not date_val or not amount_val:
                return None
            
            try:
                amount = self._parse_amount(amount_val)
            except (ValueError, TypeError) as e:
                logger.debug(f"Failed to parse amount '{amount_val}': {e}")
                return None
            
            if amount <= 0:
                return None
            
            return {
                'transaction_date': self._parse_date(date_val),
                'amount': amount,
                'payer_name': str(payer_val).strip() if payer_val else '',
                'payer_account': str(self._get_field(row, headers, 'account_col', '')).strip(),
                'payer_inn': str(self._get_field(row, headers, 'inn_col', '')).strip(),
                'payment_purpose': str(self._get_field(row, headers, 'purpose_col', '')).strip(),
            }
        except Exception as e:
            logger.warning(f"Error extracting transaction: {e}")
            return None
    
    def _get_field(self, row: Dict[str, Any], headers: Dict[str, str], field_type: str, default: Any = None) -> Any:
        """Получить значение поля по паттерну банка"""
        try:
            if not self.bank_pattern:
                for bank_pattern in self.BANK_PATTERNS.values():
                    for pattern in bank_pattern.get(field_type, []):
                        for header, value in row.items():
                            if pattern.lower() in header.lower():
                                return value
            
            for pattern in self.bank_pattern.get(field_type, []):
                for header, value in row.items():
                    if pattern.lower() in header.lower():
                        return value
            
            return default
        except Exception as e:
            logger.warning(f"Error getting field {field_type}: {e}")
            return default
    
    def _parse_date(self, value: Any) -> Optional[datetime.date]:
        """Парсинг даты из разных форматов"""
        if not value:
            return date.today()
            
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            formats = [
                '%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d',
                '%d/%m/%Y', '%m/%d/%Y', '%Y.%m.%d',
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        return date.today()
    
    def _parse_amount(self, value: Any) -> Decimal:
        """Парсинг суммы из разных форматов"""
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            if isinstance(value, str):
                cleaned = value.replace(' ', '').replace(',', '.')
                cleaned = re.sub(r'[^\d.-]', '', cleaned)
                return Decimal(cleaned)
            return Decimal('0')
        except Exception as e:
            logger.warning(f"Error parsing amount '{value}': {e}")
            return Decimal('0')

    def _parse_pdf_table(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Парсинг табличных данных из PDF (более точный метод).
        """
        transactions = []
        logger.info("Extracting tables from PDF...")
        
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                logger.info(f"PDF has {len(pdf.pages)} pages")
                
                for page_num, page in enumerate(pdf.pages, 1):
                    logger.debug(f"Processing page {page_num}")
                    
                    # Извлекаем таблицы
                    tables = page.extract_tables()
                    logger.debug(f"Found {len(tables)} tables on page {page_num}")
                    
                    for table_idx, table in enumerate(tables):
                        if not table:
                            continue
                        
                        logger.debug(f"Processing table {table_idx + 1}, {len(table)} rows")
                        
                        # Ищем заголовки
                        headers = []
                        data_start_row = 0

                        for row_idx, row in enumerate(table):
                            if not row:
                                continue
                            
                            # Проверяем, является ли строка заголовком
                            row_text = ' '.join([str(c) for c in row if c])
                            if any(keyword in row_text.lower() for keyword in ['дата', 'сумма', 'плательщик', 'назначение']):
                                headers = [str(h).lower().strip() if h else '' for h in row]
                                data_start_row = row_idx + 1
                                logger.debug(f"Found headers at row {row_idx}: {headers}")
                                break
                            
                        # Если заголовки не найдены, пробуем парсить все строки
                        if not headers:
                            for row in table[1:]:  # Пропускаем первую строку
                                transaction = self._parse_table_row(row, [])
                                if transaction:
                                    transactions.append(transaction)
                        else:
                            for row in table[data_start_row:]:
                                transaction = self._parse_table_row(row, headers)
                                if transaction:
                                    transactions.append(transaction)
                    
                    # Если таблицы не найдены, пробуем извлечь текст
                    if not transactions:
                        text = page.extract_text()
                        if text:
                            logger.debug(f'Text on page {page_num}: {text[:500]}')
                            
        except ImportError:
            logger.warning('pdfplumber not installed')
        except Exception as e:
            logger.error(f'Error extracting tables: {e}\n{traceback.format_exc()}')

        logger.info(f"Extracted {len(transactions)} transactions from PDF tables")
        return transactions


    def _parse_table_row(self, row: List, headers: List[str]) -> Optional[Dict[str, Any]]:
        """Парсинг одной строки таблицы"""
        try:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                return None

            # Очищаем значения
            values = [str(c).strip() if c else '' for c in row]
            row_text = ' '.join(values)

            # Ищем дату
            date_pattern = r'(\d{2}\.\d{2}\.\d{4})'
            date_match = re.search(date_pattern, row_text)
            if not date_match:
                return None

            date_str = date_match.group(1)

            # Ищем сумму (число с точкой или запятой)
            amount_pattern = r'([\d\s]+[.,]\d{2})'
            amount_match = re.search(amount_pattern, row_text)
            if not amount_match:
                return None

            try:
                amount_str = amount_match.group(1).replace(' ', '').replace(',', '.')
                amount = Decimal(amount_str)
            except Exception as e:
                logger.debug(f"Failed to parse amount: {e}")
                return None

            if amount <= 0:
                return None

            # Определяем плательщика и назначение
            payer_name = ''
            payment_purpose = ''
            
            if headers:
                # Пытаемся найти по заголовкам
                for i, header in enumerate(headers):
                    if i < len(values):
                        if 'плательщик' in header or 'контрагент' in header or 'отправитель' in header:
                            payer_name = values[i]
                        elif 'назначение' in header or 'purpose' in header:
                            payment_purpose = values[i]

                if not payer_name:
                    # Берём текст между датой и суммой
                    parts = re.split(r'\s{2,}|\t|\|', row_text)
                    for part in parts:
                        if part != date_str and not re.match(r'^[\d\s.,]+$', part):
                            payer_name = part
                            break
            else:
                # Извлекаем плательщика из текста
                parts = re.split(r'\s{2,}|\t|\|', row_text)
                for part in parts:
                    clean_part = part.strip()
                    if clean_part and clean_part != date_str and not re.match(r'^[\d\s.,]+$', clean_part):
                        if not payer_name:
                            payer_name = clean_part
                        else:
                            payment_purpose += ' ' + clean_part

            return {
                'transaction_date': datetime.strptime(date_str, '%d.%m.%Y').date(),
                'amount': amount,
                'payer_name': payer_name.strip()[:200],
                'payment_purpose': payment_purpose.strip()[:500],
            }
        except Exception as e:
            logger.warning(f"Error parsing table row: {e}")
            return None


class PaymentMatcher:
    """Сопоставление банковских транзакций с владельцами и начислениями"""
    
    def __init__(self):
        logger.info("Initializing PaymentMatcher")
        try:
            from users.models import Owner, ContactInfo
            from land.models import LandPlot
            self.Owner = Owner
            self.ContactInfo = ContactInfo
            self.LandPlot = LandPlot
            logger.debug("PaymentMatcher initialized successfully")
        except Exception as e:
            logger.error(f"Error initializing PaymentMatcher: {e}\n{traceback.format_exc()}")
            raise
    
    def match_owner(self, transaction: Dict[str, Any]) -> Optional[Tuple[Any, float]]:
        """
        Поиск владельца по транзакции.
        """
        logger.info(f"Matching owner for transaction: {transaction.get('payer_name', 'Unknown')}")
        
        try:
            payer_name = transaction.get('payer_name', '').lower()
            payment_purpose = transaction.get('payment_purpose', '').lower()
            matched_uid = transaction.get('matched_uid', '')
            
            # 1. САМЫЙ ПРИОРИТЕТНЫЙ: Поиск по UID из квитанции
            if matched_uid:
                from .models import Assessment
                try:
                    assessment = Assessment.objects.select_related('owner').get(payment_uid=matched_uid)
                    logger.info(f"Owner found by UID {matched_uid}: {assessment.owner.full_name} (confidence: 100%)")
                    return (assessment.owner, 100.0)
                except Assessment.DoesNotExist:
                    logger.warning(f"Assessment with UID {matched_uid} not found")
                except Exception as e:
                    logger.error(f"Error searching by UID: {e}")
            
            # 2. Поиск по UID в назначении платежа
            snt_id_pattern = r'SNT-(\d{6})'
            snt_match = re.search(snt_id_pattern, payment_purpose, re.IGNORECASE)
            if snt_match:
                assessment_id = int(snt_match.group(1))
                from .models import Assessment
                try:
                    assessment = Assessment.objects.select_related('owner').get(id=assessment_id)
                    logger.info(f"Owner found by assessment ID {assessment_id}: {assessment.owner.full_name} (confidence: 100%)")
                    return (assessment.owner, 100.0)
                except Assessment.DoesNotExist:
                    logger.warning(f"Assessment with ID {assessment_id} not found")
                except Exception as e:
                    logger.error(f"Error searching by assessment ID: {e}")
                
            # 3. Поиск по номеру участка
            plot_pattern = r'уч\.?№?(\d+)'
            plot_match = re.search(plot_pattern, payment_purpose, re.IGNORECASE)
            if plot_match:
                plot_number = plot_match.group(1)
                from land.models import LandPlot
                try:
                    plot = LandPlot.objects.filter(plot_number=plot_number).first()
                    if plot and plot.owners.exists():
                        owner = plot.owners.first()
                        logger.info(f"Owner found by plot number {plot_number}: {owner.full_name} (confidence: 95%)")
                        return (owner, 95.0)
                except Exception as e:
                    logger.warning(f"Error searching by plot number: {e}")
            
            # 4. Поиск по ФИО
            if payer_name and len(payer_name) > 3:
                try:
                    owners = self.Owner.objects.all()
                    best_match = None
                    best_score = 0
                    
                    for owner in owners:
                        owner_name_lower = owner.full_name.lower()
                        # Проверяем совпадение ФИО
                        if payer_name == owner_name_lower:
                            logger.info(f"Exact match found: {owner.full_name} (confidence: 90%)")
                            return (owner, 90.0)
                        
                        # Проверяем частичное совпадение
                        if payer_name in owner_name_lower and len(payer_name) > 10:
                            score = 70
                            if score > best_score:
                                best_score = score
                                best_match = owner
                    
                    if best_match:
                        logger.info(f"Partial match found: {best_match.full_name} (confidence: {best_score}%)")
                        return (best_match, best_score)
                        
                except Exception as e:
                    logger.error(f"Error searching by name: {e}")
            
            logger.warning(f"Failed to find owner for payment: payer={payer_name}, purpose={payment_purpose[:100]}")
            return None
            
        except Exception as e:
            logger.error(f"Error in match_owner: {e}\n{traceback.format_exc()}")
            return None
    
    def match_assessment(self, owner: Any, amount: Decimal, payment_purpose: str) -> Optional[Any]:
        """Поиск подходящего начисления для платежа"""
        logger.info(f"Matching assessment for owner {owner.full_name}, amount {amount}")
        
        try:
            from .models import Assessment
            
            # Проверяем уникальный ID
            snt_id_pattern = r'SNT-(\d{6})'
            snt_match = re.search(snt_id_pattern, payment_purpose, re.IGNORECASE)
            if snt_match:
                assessment_id = int(snt_match.group(1))
                try:
                    assessment = Assessment.objects.get(
                        id=assessment_id,
                        owner=owner,
                        status__in=[Assessment.STATUS_PENDING, Assessment.STATUS_PARTIAL, Assessment.STATUS_OVERDUE]
                    )
                    logger.info(f"Found exact assessment by ID {assessment_id}: amount={assessment.amount}, debt={assessment.debt}")
                    return assessment
                except Assessment.DoesNotExist:
                    logger.debug(f"No assessment found with ID {assessment_id}")
            
            # Ищем неоплаченные начисления
            assessments = Assessment.objects.filter(
                owner=owner,
                status__in=[Assessment.STATUS_PENDING, Assessment.STATUS_PARTIAL, Assessment.STATUS_OVERDUE]
            ).order_by('period__due_date')
            
            logger.debug(f"Found {assessments.count()} unpaid assessments for owner")
            
            exact_match = assessments.filter(amount=amount).first()
            if exact_match:
                logger.info(f"Found exact amount match: assessment {exact_match.id}, amount={exact_match.amount}")
                return exact_match
            
            for assessment in assessments:
                if assessment.debt >= amount:
                    logger.info(f"Found assessment with sufficient debt: {assessment.id}, debt={assessment.debt}, amount={amount}")
                    return assessment
            
            if assessments.first():
                logger.warning(f"No suitable assessment found, returning first: {assessments.first().id}")
                return assessments.first()
            
            logger.warning(f"No assessments found for owner {owner.full_name}")
            return None
            
        except Exception as e:
            logger.error(f"Error in match_assessment: {e}\n{traceback.format_exc()}")
            return None
    
    def process_and_update_payments(self, transaction: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Обработка транзакции с автоматическим обновлением статуса начисления.
        """
        logger.info("=" * 50)
        logger.info(f"Processing payment transaction: {transaction.get('payer_name', 'Unknown')}")
        
        from .models import Assessment, Payment
        from decimal import Decimal

        result = {
            'matched': False,
            'payment_created': False,
            'assessment_updated': False,
            'message': ''
        }

        try:
            amount = transaction.get('amount', Decimal('0'))
            payment_purpose = transaction.get('payment_purpose', '')
            payer_name = transaction.get('payer_name', '')

            logger.info(f"Transaction details: amount={amount}, payer={payer_name[:50]}, purpose={payment_purpose[:100]}")

            # 1. Поиск владельца
            logger.info("Step 1: Finding owner...")
            owner_match = self.match_owner(transaction)
            if not owner_match:
                result['message'] = f'Не найден владелец для {payer_name}'
                logger.warning(result['message'])
                return result

            owner, confidence = owner_match
            result['matched_owner'] = owner.full_name
            result['confidence'] = confidence
            logger.info(f"Owner found: {owner.full_name} (confidence: {confidence}%)")

            # 2. Поиск начисления
            logger.info("Step 2: Finding assessment...")
            assessment = self.match_assessment(owner, amount, payment_purpose)
            if not assessment:
                result['message'] = f'Не найдено подходящее начисление для {owner.full_name}'
                logger.warning(result['message'])
                return result

            result['matched_assessment_id'] = assessment.id
            result['matched_assessment_amount'] = str(assessment.amount)
            result['current_debt'] = str(assessment.debt)
            logger.info(f"Assessment found: ID={assessment.id}, amount={assessment.amount}, debt={assessment.debt}")

            # 3. Создаём уникальный transaction_id, если его нет
            transaction_id = transaction.get('transaction_id', '')
            if not transaction_id:
                transaction_id = f"MANUAL_{datetime.now().strftime('%Y%m%d%H%M%S')}_{assessment.id}"
                logger.info(f"Generated transaction ID: {transaction_id}")
            
            # Проверяем существующие платежи
            logger.info("Step 3: Checking for duplicate payments...")
            existing_payment = Payment.objects.filter(transaction_id=transaction_id).first()
            if existing_payment:
                result['matched'] = True
                result['payment_exists'] = True
                result['payment_id'] = existing_payment.id
                result['message'] = f'Платёж уже обработан (ID: {existing_payment.id})'
                logger.info(result['message'])
                return result

            # Также проверяем по UID и сумме
            matched_uid = transaction.get('matched_uid', '')
            if matched_uid:
                existing_payment = Payment.objects.filter(
                    matched_uid=matched_uid,
                    amount=amount
                ).first()
                if existing_payment:
                    result['matched'] = True
                    result['payment_exists'] = True
                    result['payment_id'] = existing_payment.id
                    result['message'] = f'Платёж с UID {matched_uid} уже обработан'
                    logger.info(result['message'])
                    return result

            # 4. Создаём платёж
            logger.info("Step 4: Creating payment record...")
            try:
                payment = Payment.objects.create(
                    assessment=assessment,
                    amount=amount,
                    payment_date=transaction.get('transaction_date', date.today()),
                    payment_method='bank',
                    bank_name=transaction.get('bank_name', ''),
                    bank_account=transaction.get('payer_account', ''),
                    transaction_id=transaction_id,
                    matched_uid=matched_uid,
                    payment_purpose=payment_purpose[:500],
                    status=Payment.STATUS_PROCESSED,
                )
                logger.info(f"Payment created: ID={payment.id}, amount={amount}")
            except Exception as e:
                logger.error(f"Error creating payment: {e}")
                result['message'] = f'Ошибка создания платежа: {str(e)}'
                return result

            result['payment_created'] = True
            result['payment_id'] = payment.id
            result['payment_amount'] = str(amount)

            # 5. Проверяем статус начисления после оплаты
            logger.info("Step 5: Updating assessment status...")
            assessment.refresh_from_db()
            result['new_debt'] = str(assessment.debt)
            result['assessment_status'] = assessment.get_status_display()

            if assessment.status == Assessment.STATUS_PAID:
                result['message'] = f'✅ Начисление полностью оплачено!'
                logger.info(f"Assessment {assessment.id} fully paid")
            else:
                result['message'] = f'💰 Внесён платёж {amount} ₽. Остаток долга: {assessment.debt} ₽'
                logger.info(f"Payment applied, remaining debt: {assessment.debt}")

            result['matched'] = True
            logger.info(f"Transaction processed successfully: {result['message']}")
            
        except Exception as e:
            logger.error(f"Error processing payment: {e}\n{traceback.format_exc()}")
            result['message'] = f'Ошибка обработки: {str(e)}'
            
        return result